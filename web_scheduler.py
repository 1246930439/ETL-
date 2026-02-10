from flask import Flask, render_template, request, jsonify, redirect, url_for, session
from flask_mail import Mail, Message
from werkzeug.security import check_password_hash, generate_password_hash
import sqlite3
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import threading
import subprocess
import os
import sys
import json
import threading
from datetime import datetime, timedelta, timezone
import time
import pymysql
import hashlib
import logging
import secrets
from functools import wraps
from croniter import croniter
import smtplib

try:
    import psycopg2
except ImportError:
    psycopg2 = None

try:
    import pandas as pd
except ImportError:
    pd = None

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:
    openpyxl = None
    get_column_letter = None

import tempfile
import os

# 用于防止任务重复执行的全局变量
executing_tasks = set()
executing_tasks_lock = threading.Lock()
executing_sql_alerts = set()
executing_sql_alerts_lock = threading.Lock()

# 日志收集器类
class LogCollector:
    """用于收集SQL预警执行过程中的日志信息"""
    def __init__(self):
        self.logs = []
        self.original_info = None
        self.original_error = None
        self.original_warning = None
    
    def start_collecting(self):
        """开始收集日志"""
        # 保存原始的日志方法
        self.original_info = logger.info
        self.original_error = logger.error
        self.original_warning = logger.warning
        
        # 替换为自定义的日志方法
        logger.info = self.collect_info
        logger.error = self.collect_error
        logger.warning = self.collect_warning
    
    def stop_collecting(self):
        """停止收集日志并恢复原始方法"""
        if self.original_info:
            logger.info = self.original_info
        if self.original_error:
            logger.error = self.original_error
        if self.original_warning:
            logger.warning = self.original_warning
    
    def collect_info(self, msg, *args, **kwargs):
        """收集info级别的日志"""
        formatted_msg = msg % args if args else msg
        self.logs.append(f"INFO: {formatted_msg}")
        self.original_info(msg, *args, **kwargs)
    
    def collect_error(self, msg, *args, **kwargs):
        """收集error级别的日志"""
        formatted_msg = msg % args if args else msg
        self.logs.append(f"ERROR: {formatted_msg}")
        self.original_error(msg, *args, **kwargs)
    
    def collect_warning(self, msg, *args, **kwargs):
        """收集warning级别的日志"""
        formatted_msg = msg % args if args else msg
        self.logs.append(f"WARNING: {formatted_msg}")
        self.original_warning(msg, *args, **kwargs)
    
    def get_collected_logs(self):
        """获取收集到的日志"""
        return "\n".join(self.logs)

app = Flask(__name__)
app.secret_key = 'your-secret-key-here'  # 在生产环境中应该使用更强的密钥

# 配置会话
app.config['SESSION_COOKIE_SECURE'] = False  # 在开发环境中允许HTTP
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'

# 配置JSON处理
app.config['JSON_SORT_KEYS'] = False

# 禁用严格的Content-Type检查
app.config['TRAP_BAD_REQUEST_ERRORS'] = True
app.config['TRAP_HTTP_EXCEPTIONS'] = True

# 创建中间件来处理GET请求带Content-Type的问题
class ContentTypeMiddleware:
    def __init__(self, app):
        self.app = app
    
    def __call__(self, environ, start_response):
        # 如果是GET请求且Content-Type是application/json，移除它
        if environ.get('REQUEST_METHOD', '') == 'GET' and environ.get('CONTENT_TYPE', '') == 'application/json':
            environ = dict(environ)  # 创建副本
            environ['CONTENT_TYPE'] = ''
        
        return self.app(environ, start_response)

# 应用中间件
app.wsgi_app = ContentTypeMiddleware(app.wsgi_app)

# 自定义错误处理器
@app.errorhandler(400)
def bad_request(error):
    """处理400错误"""
    logger.error(f"400错误: {error}")
    logger.error(f"错误详情: {str(error.description) if hasattr(error, 'description') else '无详细信息'}")
    
    # 如果是GET请求且路径是/api/notification-logs，尝试直接处理
    if request.method == 'GET' and request.path == '/api/notification-logs':
        logger.warning("GET请求到/api/notification-logs出现400错误，尝试直接处理")
        try:
            # 直接执行API逻辑，绕过Flask的请求处理
            import sqlite3
            
            # 获取分页参数
            page_str = request.args.get('page', '1')
            per_page_str = request.args.get('per_page', '20')
            task_id_str = request.args.get('task_id')
            task_name = request.args.get('task_name')
            alert_type = request.args.get('alert_type')
            status = request.args.get('status')
            
            # 转换参数类型
            try:
                page = int(page_str) if page_str else 1
                per_page = int(per_page_str) if per_page_str else 20
                task_id = int(task_id_str) if task_id_str else None
            except ValueError as e:
                logger.error(f"Parameter conversion error: {str(e)}")
                return jsonify({'error': f'Invalid parameter format: {str(e)}'}), 400
            
            # 参数验证
            if page < 1:
                logger.error(f"Invalid page parameter: {page}")
                return jsonify({'error': 'Invalid page parameter'}), 400
            
            if per_page < 1 or per_page > 100:
                logger.error(f"Invalid per_page parameter: {per_page}")
                return jsonify({'error': 'Invalid per_page parameter'}), 400
            
            # 限制每页最大记录数
            per_page = min(per_page, 100)

            # 连接数据库
            conn = sqlite3.connect('scheduler.db')
            cursor = conn.cursor()

            # 构建查询条件
            where_conditions = []
            params = []

            if task_id:
                where_conditions.append("nl.task_id = ?")
                params.append(task_id)

            if task_name:
                where_conditions.append("nl.task_name LIKE ?")
                params.append(f"%{task_name}%")

            if alert_type:
                where_conditions.append("nl.alert_type = ?")
                params.append(alert_type)

            if status:
                where_conditions.append("nl.status = ?")
                params.append(status)

            where_clause = "WHERE " + " AND ".join(where_conditions) if where_conditions else ""

            # 获取总记录数
            cursor.execute(f'''
                SELECT COUNT(*)
                FROM notification_logs nl
                {where_clause}
            ''', params)
            total_logs = cursor.fetchone()[0]

            # 计算偏移量
            offset = (page - 1) * per_page

            # 获取日志记录
            cursor.execute(f'''
                SELECT nl.id, nl.task_id, nl.task_name, nl.alert_type, 
                       ec.config_name, nl.recipients, nl.subject, nl.status, 
                       nl.error_message, nl.sent_time
                FROM notification_logs nl
                LEFT JOIN email_configs ec ON nl.email_config_id = ec.id
                {where_clause}
                ORDER BY nl.sent_time DESC
                LIMIT ? OFFSET ?
            ''', params + [per_page, offset])

            logs = cursor.fetchall()
            conn.close()
            
            # 转换为字典列表
            logs_list = []
            for log in logs:
                log_id, task_id, task_name, alert_type, config_name, recipients, subject, status, error_message, sent_time = log
                
                logs_list.append({
                    'id': log_id,
                    'task_id': task_id,
                    'task_name': task_name,
                    'alert_type': alert_type,
                    'config_name': config_name,
                    'recipients': recipients,
                    'subject': subject,
                    'status': status,
                    'error_message': error_message,
                    'sent_time': sent_time
                })

            result = {
                'logs': logs_list,
                'pagination': {
                    'page': page,
                    'per_page': per_page,
                    'total': total_logs,
                    'pages': (total_logs + per_page - 1) // per_page
                }
            }
            
            logger.info(f"成功处理400错误，返回{len(logs_list)}条记录")
            return jsonify(result)
            
        except Exception as e:
            logger.error(f"直接处理API请求失败: {str(e)}")
            import traceback
            traceback.print_exc()
    
    # 尝试获取更详细的错误信息
    error_message = "请求错误"
    if hasattr(error, 'description') and error.description:
        error_message = str(error.description)
    
    return jsonify({'error': error_message, 'detail': str(error)}), 400

# 邮件配置
app.config['MAIL_SERVER'] = 'smtp.gmail.com'  # 默认SMTP服务器
app.config['MAIL_PORT'] = 587  # 默认端口
app.config['MAIL_USE_TLS'] = True  # 启用TLS
app.config['MAIL_USERNAME'] = ''  # 邮箱账号
app.config['MAIL_PASSWORD'] = ''  # 邮箱密码
app.config['MAIL_DEFAULT_SENDER'] = ''  # 默认发件人
app.config['MAIL_NOTIFICATION_ENABLED'] = False  # 是否启用邮件通知

mail = Mail(app)

# 配置日志
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s %(levelname)s %(name)s %(message)s',
    handlers=[
        logging.FileHandler("scheduler.log"),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)

# 添加详细的请求/响应日志记录 (临时禁用以排查问题)
# @app.before_request
# def log_request():
#     """记录每个请求的详细信息"""
#     # 只记录API请求以避免日志过多
#     if request.path.startswith('/api/') or request.path == '/login':
#         logger.info(f"=== REQUEST START ===")
#         logger.info(f"Method: {request.method}")
#         logger.info(f"Path: {request.path}")
#         logger.info(f"Headers: {dict(request.headers)}")
#         logger.info(f"Args: {dict(request.args)}")
#         logger.info(f"Form: {dict(request.form)}")
#         logger.info(f"JSON: {request.get_json(silent=True)}")
#         logger.info(f"Session: {dict(session)}")
#         logger.info(f"Remote Address: {request.remote_addr}")
#         logger.info(f"User Agent: {request.headers.get('User-Agent', 'Unknown')}")

# @app.after_request
# def log_response(response):
#     """记录每个响应的详细信息"""
#     # 只记录API请求以避免日志过多
#     if request.path.startswith('/api/') or request.path == '/login':
#         logger.info(f"Response Status: {response.status_code}")
#         logger.info(f"Response Headers: {dict(response.headers)}")
#         # 只记录小响应的内容
#         if response.content_length and response.content_length < 1000:
#             logger.info(f"Response Data: {response.get_data(as_text=True)}")
#         logger.info(f"=== REQUEST END ===")
#     return response

# @app.errorhandler(Exception)
# def log_exception(e):
#     """记录未处理的异常"""
#     logger.error(f"Unhandled Exception: {str(e)}", exc_info=True)
#     return jsonify({'error': 'Internal server error'}), 500

# 数据库文件路径
DB_PATH = 'scheduler.db'

# 存储正在运行的任务线程
running_tasks = {}


def init_db():
    """初始化数据库"""
    logger.info("Initializing database")
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()

    # 检查是否需要迁移旧的tasks表
    try:
        cursor.execute("SELECT task_type FROM tasks LIMIT 1")
        # 如果没有异常，说明新表结构已经存在
        new_structure_exists = True
    except sqlite3.OperationalError:
        # 如果出现异常，说明是旧表结构
        new_structure_exists = False

    if not new_structure_exists:
        # 检查是否存在旧的tasks表
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='tasks'")
        old_table_exists = cursor.fetchone()

        if old_table_exists:
            # 重命名旧表
            cursor.execute("ALTER TABLE tasks RENAME TO tasks_old")

        # 创建新的任务表（增加task_type字段来区分Python脚本任务和SQL脚本任务）
        cursor.execute('''
            CREATE TABLE tasks (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                task_type TEXT NOT NULL,  -- 'python' 或 'sql'
                script_path TEXT,         -- Python脚本路径
                sql_script_id INTEGER,    -- SQL脚本ID
                schedule_interval INTEGER NOT NULL,  -- 执行间隔（秒）
                last_run TIMESTAMP,
                next_run TIMESTAMP,
                is_active BOOLEAN DEFAULT 1,
                dependencies TEXT,         -- 依赖任务ID列表，逗号分隔
                max_retries INTEGER DEFAULT 0,  -- 最大重试次数
                retry_delay INTEGER DEFAULT 60  -- 重试延迟（秒）
            )
        ''')

        # 如果存在旧数据，迁移数据
        if old_table_exists:
            cursor.execute('''
                INSERT INTO tasks (id, name, task_type, script_path, schedule_interval, last_run, next_run, is_active, dependencies)
                SELECT id, name, 'python', script_path, schedule_interval, last_run, next_run, is_active, dependencies
                FROM tasks_old
            ''')
            # 删除旧表
            cursor.execute("DROP TABLE tasks_old")
    else:
        # 检查是否需要添加依赖关系字段
        cursor.execute("PRAGMA table_info(tasks)")
        columns = [column[1] for column in cursor.fetchall()]

        # 如果不存在dependencies字段，则添加
        if 'dependencies' not in columns:
            logger.info("Adding dependencies column to tasks table")
            try:
                # SQLite添加列
                cursor.execute("ALTER TABLE tasks ADD COLUMN dependencies TEXT")
            except sqlite3.OperationalError as e:
                logger.warning(f"Could not add dependencies column: {e}")

        # 如果不存在max_retries字段，则添加
        if 'max_retries' not in columns:
            logger.info("Adding max_retries column to tasks table")
            try:
                cursor.execute("ALTER TABLE tasks ADD COLUMN max_retries INTEGER DEFAULT 0")
            except sqlite3.OperationalError as e:
                logger.warning(f"Could not add max_retries column: {e}")

        # 如果不存在retry_delay字段，则添加
        if 'retry_delay' not in columns:
            logger.info("Adding retry_delay column to tasks table")
            try:
                cursor.execute("ALTER TABLE tasks ADD COLUMN retry_delay INTEGER DEFAULT 60")
            except sqlite3.OperationalError as e:
                logger.warning(f"Could not add retry_delay column: {e}")

        # 如果不存在cron_expression字段，则添加
        if 'cron_expression' not in columns:
            logger.info("Adding cron_expression column to tasks table")
            try:
                cursor.execute("ALTER TABLE tasks ADD COLUMN cron_expression TEXT")
            except sqlite3.OperationalError as e:
                logger.warning(f"Could not add cron_expression column: {e}")

    # 检查db_configs表是否存在database字段，如果存在则需要重建表
    cursor.execute("PRAGMA table_info(db_configs)")
    columns = [column[1] for column in cursor.fetchall()]

    if 'database' in columns:
        # 如果存在database字段，则重建表
        logger.info("Removing database column from db_configs table")
        # 创建新表（不含database字段）
        cursor.execute('''
            CREATE TABLE db_configs_new (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                db_type TEXT NOT NULL,
                host TEXT,
                port INTEGER,
                username TEXT,
                password TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')

        # 复制数据
        cursor.execute('''
            INSERT INTO db_configs_new (id, name, db_type, host, port, username, password, created_at)
            SELECT id, name, db_type, host, port, username, password, created_at FROM db_configs
        ''')

        # 删除旧表并重命名新表
        cursor.execute('DROP TABLE db_configs')
        cursor.execute('ALTER TABLE db_configs_new RENAME TO db_configs')
    else:
        # 如果不存在database字段，正常创建表
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS db_configs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                db_type TEXT NOT NULL,  -- mysql, sqlite, postgresql等
                host TEXT,
                port INTEGER,
                username TEXT,
                password TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')

    # 创建SQL脚本表（不包含database_name字段）
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS sql_scripts (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            db_config_id INTEGER,
            sql_content TEXT NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (db_config_id) REFERENCES db_configs (id)
        )
    ''')

    # 检查是否需要移除旧的database_name字段
    cursor.execute("PRAGMA table_info(sql_scripts)")
    columns = [column[1] for column in cursor.fetchall()]
    if 'database_name' in columns:
        logger.info("Removing database_name column from sql_scripts table")
        # SQLite不支持直接删除列，需要重建表
        cursor.execute('''
            CREATE TABLE sql_scripts_new (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                db_config_id INTEGER,
                sql_content TEXT NOT NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (db_config_id) REFERENCES db_configs (id)
            )
        ''')

        # 复制数据
        cursor.execute('''
            INSERT INTO sql_scripts_new (id, name, db_config_id, sql_content, created_at)
            SELECT id, name, db_config_id, sql_content, created_at FROM sql_scripts
        ''')

        # 删除旧表并重命名新表
        cursor.execute('DROP TABLE sql_scripts')
        cursor.execute('ALTER TABLE sql_scripts_new RENAME TO sql_scripts')

    # 创建用户表
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')

    # 创建用户API Tokens表（修复可能的表结构问题）
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS user_tokens (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            token TEXT UNIQUE NOT NULL,
            name TEXT NOT NULL,
            expires_at TIMESTAMP NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (user_id) REFERENCES users (id)
        )
    ''')

    # 创建任务日志表
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS task_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            task_id INTEGER NOT NULL,
            status TEXT NOT NULL,  -- 'success' 或 'failed'
            message TEXT,
            execution_time TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (task_id) REFERENCES tasks (id)
        )
    ''')

    # 创建邮件配置表
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS email_configs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            config_name TEXT NOT NULL,
            mail_server TEXT NOT NULL,
            mail_port INTEGER NOT NULL,
            mail_use_tls BOOLEAN DEFAULT 1,
            mail_username TEXT NOT NULL,
            mail_password TEXT NOT NULL,
            mail_default_sender TEXT,
            mail_notification_enabled BOOLEAN DEFAULT 0,
            is_default BOOLEAN DEFAULT 0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    # 创建任务状态预警表
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS task_alerts (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            task_id INTEGER,  -- 关联的任务ID，可以为空表示全局预警
            alert_type TEXT NOT NULL,  -- 'success' 或 'failure'
            email_config_id INTEGER NOT NULL,
            recipients TEXT NOT NULL,   -- 收件人邮箱列表，逗号分隔
            is_enabled BOOLEAN DEFAULT 1,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (task_id) REFERENCES tasks(id) ON DELETE SET NULL,
            FOREIGN KEY (email_config_id) REFERENCES email_configs(id)
        )
    ''')
    
    # 创建预警日志表
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS notification_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            task_id INTEGER NOT NULL,
            task_name TEXT NOT NULL,
            alert_type TEXT NOT NULL,  -- 'success' 或 'failure'
            email_config_id INTEGER,
            recipients TEXT,            -- 收件人邮箱列表，逗号分隔
            subject TEXT,              -- 邮件主题
            body TEXT,                 -- 邮件内容
            status TEXT NOT NULL,      -- 'sent' 或 'failed'
            error_message TEXT,        -- 发送失败时的错误信息
            sent_time TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (task_id) REFERENCES tasks (id),
            FOREIGN KEY (email_config_id) REFERENCES email_configs(id)
        )
    ''')
    
    # 检查task_alerts表是否需要添加task_id字段
    cursor.execute("PRAGMA table_info(task_alerts)")
    task_alert_columns = [column[1] for column in cursor.fetchall()]
    
    # 如果task_id字段不存在，添加它
    if 'task_id' not in task_alert_columns:
        logger.info("Adding task_id column to task_alerts table")
        cursor.execute("ALTER TABLE task_alerts ADD COLUMN task_id INTEGER")
        logger.info("Added task_id column to task_alerts table")
    else:
        logger.info("Task alerts table already has task_id field")
    
    # 如果name字段不存在，添加它
    if 'name' not in task_alert_columns:
        logger.info("Adding name column to task_alerts table")
        cursor.execute("ALTER TABLE task_alerts ADD COLUMN name TEXT")
        logger.info("Added name column to task_alerts table")
    else:
        logger.info("Task alerts table already has name field")
    
    # 创建SQL查询预警表
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS sql_alerts (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            sql_script_id INTEGER NOT NULL,
            email_config_id INTEGER NOT NULL,
            recipients TEXT NOT NULL,   -- 收件人邮箱列表，逗号分隔
            condition_type TEXT DEFAULT 'not_empty',  -- 'not_empty' 或其他条件
            threshold INTEGER DEFAULT 1,  -- 阈值，用于条件判断
            is_enabled BOOLEAN DEFAULT 1,
            last_check TIMESTAMP,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (sql_script_id) REFERENCES sql_scripts(id) ON DELETE CASCADE,
            FOREIGN KEY (email_config_id) REFERENCES email_configs(id)
        )
    ''')
    
    # 创建SQL预警日志表
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS sql_alert_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            alert_id INTEGER NOT NULL,
            alert_name TEXT NOT NULL,
            status TEXT NOT NULL,  -- 'success', 'failed', 'triggered', 'no_trigger'
            message TEXT,
            execution_time TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (alert_id) REFERENCES sql_alerts (id)
        )
    ''')
    
    # 检查sql_alerts表是否有threshold字段，如果没有则添加
    cursor.execute("PRAGMA table_info(sql_alerts)")
    sql_alert_columns = [column[1] for column in cursor.fetchall()]
    
    if 'threshold' not in sql_alert_columns:
        logger.info("Adding threshold column to sql_alerts table")
        cursor.execute("ALTER TABLE sql_alerts ADD COLUMN threshold INTEGER DEFAULT 1")
        logger.info("Added threshold column to sql_alerts table")
    
    # 检查sql_alerts表是否有cron_expression字段，如果没有则添加
    if 'cron_expression' not in sql_alert_columns:
        logger.info("Adding cron_expression column to sql_alerts table")
        cursor.execute("ALTER TABLE sql_alerts ADD COLUMN cron_expression TEXT")
        logger.info("Added cron_expression column to sql_alerts table")
    
    # 检查sql_alerts表是否有next_check字段，如果没有则添加
    if 'next_check' not in sql_alert_columns:
        logger.info("Adding next_check column to sql_alerts table")
        cursor.execute("ALTER TABLE sql_alerts ADD COLUMN next_check TIMESTAMP")
        logger.info("Added next_check column to sql_alerts table")
    
    # 检查email_configs表是否有新字段，如果没有则添加
    cursor.execute("PRAGMA table_info(email_configs)")
    email_columns = [column[1] for column in cursor.fetchall()]
    
    if 'config_name' not in email_columns:
        cursor.execute("ALTER TABLE email_configs ADD COLUMN config_name TEXT NOT NULL DEFAULT 'default'")
    
    if 'is_default' not in email_columns:
        cursor.execute("ALTER TABLE email_configs ADD COLUMN is_default BOOLEAN DEFAULT 0")
        
    if 'updated_at' not in email_columns:
        cursor.execute("ALTER TABLE email_configs ADD COLUMN updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP")

    # 检查是否有用户，如果没有则创建默认管理员
    cursor.execute("SELECT COUNT(*) FROM users")
    user_count = cursor.fetchone()[0]
    
    if user_count == 0:
        # 创建默认管理员用户
        default_username = "admin"
        default_password = "admin123"  # 首次登录后请立即修改
        password_hash = generate_password_hash(default_password)
        created_at = get_beijing_time().isoformat()
        
        cursor.execute('''
            INSERT INTO users (username, password_hash, created_at)
            VALUES (?, ?, ?)
        ''', (default_username, password_hash, created_at))
        
        logger.info(f"Created default admin user: {default_username}")
    
    conn.commit()
    conn.close()
    logger.info("Database initialization completed")


# !/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
import sys
import logging
from datetime import datetime, timedelta, timezone

# 添加tzlocal库来更好地处理时区
try:
    from tzlocal import get_localzone

    HAS_TZLOCAL = True
except ImportError:
    HAS_TZLOCAL = False


def get_beijing_time():
    """获取北京时间"""
    try:
        # Python 3.12+
        utc_time = datetime.now(timezone.utc)
    except AttributeError:
        # Older Python versions
        utc_time = datetime.utcnow().replace(tzinfo=timezone.utc)

    # 转换为北京时间 (UTC+8)
    beijing_time = utc_time + timedelta(hours=8)
    return beijing_time.replace(tzinfo=None)


def execute_script(script_path):
    """执行Python脚本"""
    logger.info(f"Executing Python script: {script_path}")
    try:
        # 处理路径映射，适应不同环境
        mapped_script_path = map_script_path(script_path)

        # 检查文件是否存在
        if not os.path.exists(mapped_script_path):
            logger.error(f"Script file not found: {mapped_script_path}")
            return None

        # 获取脚本所在的目录并将其添加到PYTHONPATH
        script_dir = os.path.dirname(os.path.abspath(mapped_script_path))
        env = os.environ.copy()
        if 'PYTHONPATH' in env:
            env['PYTHONPATH'] = script_dir + os.pathsep + env['PYTHONPATH']
        else:
            env['PYTHONPATH'] = script_dir

        # 设置环境变量确保输出编码正确
        env['PYTHONIOENCODING'] = 'utf-8'

        # 使用二进制模式读取输出，避免编码问题
        process = subprocess.Popen(
            [sys.executable, mapped_script_path],
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            cwd=script_dir,  # 设置工作目录为脚本所在目录
            env=env  # 传递环境变量
        )

        # 等待进程完成，设置超时
        stdout, stderr = process.communicate(timeout=300)  # 5分钟超时

        # 手动解码输出，忽略编码错误
        try:
            stdout_decoded = stdout.decode('utf-8')
        except UnicodeDecodeError:
            stdout_decoded = stdout.decode('utf-8', errors='ignore')

        try:
            stderr_decoded = stderr.decode('utf-8')
        except UnicodeDecodeError:
            stderr_decoded = stderr.decode('utf-8', errors='ignore')

        result = subprocess.CompletedProcess(
            args=[sys.executable, mapped_script_path],
            returncode=process.returncode,
            stdout=stdout_decoded,
            stderr=stderr_decoded
        )

        logger.info(f"Script output: {result.stdout}")
        if result.stderr:
            logger.warning(f"Script errors: {result.stderr}")
        return result
    except subprocess.TimeoutExpired:
        logger.error(f"Script execution timed out: {mapped_script_path}")
        process.kill()
        stdout, stderr = process.communicate()

        # 手动解码输出，忽略编码错误
        try:
            stdout_decoded = stdout.decode('utf-8')
        except UnicodeDecodeError:
            stdout_decoded = stdout.decode('utf-8', errors='ignore')

        try:
            stderr_decoded = stderr.decode('utf-8')
        except UnicodeDecodeError:
            stderr_decoded = stderr.decode('utf-8', errors='ignore')

        return subprocess.CompletedProcess(
            args=[sys.executable, mapped_script_path],
            returncode=-1,
            stdout=stdout_decoded,
            stderr="脚本执行超时"
        )
    except Exception as e:
        logger.error(f"Error executing script {mapped_script_path}: {e}", exc_info=True)
        # 确保进程被终止
        if 'process' in locals():
            try:
                process.kill()
                process.communicate()
            except:
                pass
        return subprocess.CompletedProcess(
            args=[sys.executable, mapped_script_path],
            returncode=-1,
            stdout="",
            stderr=f"执行脚本时发生错误: {str(e)}"
        )


def normalize_script_path_for_storage(script_path):
    """
    标准化脚本路径以用于存储
    这个函数将实际路径转换为适合存储的形式，以便在不同环境中可以正确映射
    """
    # 获取当前项目根目录
    project_root = os.path.dirname(os.path.abspath(__file__))

    try:
        # 尝试获取相对于项目根目录的路径
        rel_path = os.path.relpath(script_path, project_root)
        # 如果相对路径不以 .. 开头，说明在项目目录内，可以使用相对路径存储
        if not rel_path.startswith('..'):
            return rel_path
    except ValueError:
        # 不同驱动器上的路径无法计算相对路径，继续下面的处理
        pass

    # 如果是外部绝对路径，只存储文件名
    # 在不同环境中会通过map_script_path函数在uploads目录中寻找同名文件
    filename = os.path.basename(script_path)
    uploads_dir = os.path.join('uploads', filename)
    return uploads_dir


def map_script_path(script_path):
    """
    映射脚本路径以适应不同环境
    这个函数可以根据实际部署环境转换脚本路径
    """
    # 获取当前项目根目录
    project_root = os.path.dirname(os.path.abspath(__file__))

    # 如果脚本路径是相对路径，则相对于项目根目录解析
    if not os.path.isabs(script_path):
        return os.path.join(project_root, script_path)

    # 如果是绝对路径，检查是否在项目目录内
    try:
        # 尝试获取相对于项目根目录的路径
        rel_path = os.path.relpath(script_path, project_root)
        # 如果相对路径不以 .. 开头，说明在项目目录内
        if not rel_path.startswith('..'):
            return script_path
    except ValueError:
        # 不同驱动器上的路径无法计算相对路径，继续下面的处理
        pass

    # 如果是外部绝对路径，直接返回
    # 注意：这里可能需要根据实际部署情况进行更复杂的映射逻辑
    filename = os.path.basename(script_path)
    uploads_dir = os.path.join(project_root, 'uploads')

    # 检查 uploads 目录中是否存在同名文件
    mapped_path = os.path.join(uploads_dir, filename)
    if os.path.exists(mapped_path):
        return mapped_path

    # 如果没有找到映射路径，返回原始路径
    # 这种情况下，用户需要确保脚本在目标环境中存在且路径正确
    logger.warning(f"Could not map script path {script_path}. Using original path.")
    return script_path


def execute_sql_on_database(db_config, sql_content):
    """在指定的数据库上执行SQL语句"""
    db_type = db_config['db_type']
    logger.info(f"Executing SQL on database type: {db_type}")

    try:
        if db_type == 'mysql':
            # 连接到MySQL数据库
            logger.info(f"Connecting to MySQL database: {db_config['host']}:{db_config['port']}")
            connection_params = {
                'host': db_config['host'],
                'port': db_config['port'] or 3306,
                'user': db_config['username'],
                'password': db_config['password'],
                'charset': 'utf8mb4',
                'autocommit': True,
                'cursorclass': pymysql.cursors.DictCursor,
                'read_timeout': 30,
                'write_timeout': 30
            }

            connection = pymysql.connect(**connection_params)

            with connection:
                # 按分号分割SQL语句
                sql_statements = [stmt.strip() for stmt in sql_content.split(';') if stmt.strip()]
                results = []

                with connection.cursor() as cursor:
                    for sql_stmt in sql_statements:
                        logger.info(f"Executing SQL statement: {sql_stmt[:100]}...")
                        # 执行单条SQL语句
                        cursor.execute(sql_stmt)
                        affected_rows = cursor.rowcount
                        # 如果是查询语句，获取结果
                        if sql_stmt.strip().upper().startswith('SELECT'):
                            result = cursor.fetchall()
                            results.append({'sql': sql_stmt, 'result': result, 'affected_rows': None})
                        else:
                            results.append({'sql': sql_stmt, 'result': None, 'affected_rows': affected_rows})

                logger.info("MySQL SQL execution completed successfully")
                return True, results

        elif db_type == 'postgresql':
            # 连接到PostgreSQL数据库
            if psycopg2 is None:
                logger.error("PostgreSQL driver not installed")
                return False, "PostgreSQL驱动未安装"

            logger.info(f"Connecting to PostgreSQL database: {db_config['host']}:{db_config['port']}")
            connection_params = {
                'host': db_config['host'],
                'port': db_config['port'] or 5432,
                'user': db_config['username'],
                'password': db_config['password'],
                'connect_timeout': 30
            }

            connection = psycopg2.connect(**connection_params)

            with connection:
                # 按分号分割SQL语句
                sql_statements = [stmt.strip() for stmt in sql_content.split(';') if stmt.strip()]
                results = []

                with connection.cursor() as cursor:
                    for sql_stmt in sql_statements:
                        logger.info(f"Executing SQL statement: {sql_stmt[:100]}...")
                        # 执行单条SQL语句
                        cursor.execute(sql_stmt)
                        affected_rows = cursor.rowcount
                        # 如果是查询语句，获取结果
                        if sql_stmt.strip().upper().startswith('SELECT'):
                            result = cursor.fetchall()
                            # 获取列名
                            colnames = [desc[0] for desc in cursor.description]
                            result = {'columns': colnames, 'rows': result}
                            results.append({'sql': sql_stmt, 'result': result, 'affected_rows': None})
                        else:
                            connection.commit()
                            results.append({'sql': sql_stmt, 'result': None, 'affected_rows': affected_rows})

                logger.info("PostgreSQL SQL execution completed successfully")
                return True, results

        elif db_type == 'sqlite':
            # 连接到SQLite数据库
            logger.info(f"Connecting to SQLite database: {db_config.get('database', ':memory:')}")
            connection = sqlite3.connect(db_config.get('database', ':memory:'), timeout=30)
            connection.row_factory = sqlite3.Row  # 使得结果可以通过列名访问

            # 按分号分割SQL语句
            sql_statements = [stmt.strip() for stmt in sql_content.split(';') if stmt.strip()]
            results = []

            try:
                with connection:
                    for sql_stmt in sql_statements:
                        logger.info(f"Executing SQL statement: {sql_stmt[:100]}...")
                        cursor = connection.cursor()
                        # 执行单条SQL语句
                        cursor.execute(sql_stmt)
                        affected_rows = cursor.rowcount
                        # 如果是查询语句，获取结果
                        if sql_stmt.strip().upper().startswith('SELECT'):
                            result = cursor.fetchall()
                            # 转换为字典列表
                            result = [dict(row) for row in result]
                            results.append({'sql': sql_stmt, 'result': result, 'affected_rows': None})
                        else:
                            results.append({'sql': sql_stmt, 'result': None, 'affected_rows': affected_rows})

                logger.info("SQLite SQL execution completed successfully")
                return True, results
            finally:
                connection.close()

        else:
            logger.error(f"Unsupported database type: {db_type}")
            return False, f"不支持的数据库类型: {db_type}"

    except Exception as e:
        logger.error(f"Error executing SQL on {db_type} database: {e}", exc_info=True)
        # 确保数据库连接被关闭
        if 'connection' in locals():
            try:
                connection.close()
            except:
                pass
        return False, f"执行SQL时出错: {str(e)}"


def execute_sql_script(sql_script_id):
    """执行SQL脚本"""
    logger.info(f"Executing SQL script ID: {sql_script_id}")
    try:
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        
        # 如果更新了sql_script_id，检查新的SQL脚本是否有关联的数据库配置
        if sql_script_id is not None:
            cursor.execute('SELECT db_config_id FROM sql_scripts WHERE id = ?', (sql_script_id,))
            script_result = cursor.fetchone()
            if not script_result:
                conn.close()
                return jsonify({'error': 'SQL脚本不存在'}), 400
                
            db_config_id = script_result[0]
            if not db_config_id:
                conn.close()
                return jsonify({'error': 'SQL脚本没有关联的数据库配置，请先编辑SQL脚本并选择数据库配置'}), 400

        # 获取SQL脚本和数据库配置
        cursor.execute('''
            SELECT s.sql_content, c.db_type, c.host, c.port, c.username, c.password
            FROM sql_scripts s
            JOIN db_configs c ON s.db_config_id = c.id
            WHERE s.id = ?
        ''', (sql_script_id,))

        result = cursor.fetchone()
        conn.close()

        if not result:
            logger.error(f"SQL script not found: {sql_script_id}")
            print(f"SQL脚本不存在: {sql_script_id}")
            return False, "SQL脚本不存在"

        sql_content, db_type, host, port, username, password = result

        # 构造数据库配置字典
        db_config = {
            'db_type': db_type,
            'host': host,
            'port': port,
            'username': username,
            'password': password
        }

        # 执行SQL脚本
        success, execution_result = execute_sql_on_database(db_config, sql_content)
        if success:
            logger.info(f"SQL script executed successfully: {sql_script_id}")
            print(f"SQL脚本执行成功: {sql_script_id}")
            return True, execution_result
        else:
            logger.error(f"Failed to execute SQL script {sql_script_id}: {execution_result}")
            print(f"SQL脚本执行失败: {execution_result}")
            return False, execution_result
    except Exception as e:
        error_msg = f'SQL执行失败: {str(e)}'
        logger.error(f"Exception while executing SQL script {sql_script_id}: {e}", exc_info=True)
        print(error_msg)
        # 确保数据库连接被关闭
        if 'conn' in locals():
            try:
                conn.close()
            except:
                pass
        return False, error_msg


def log_task_execution(task_id, status, message):
    """记录任务执行日志"""
    logger.info(f"Logging task execution: task_id={task_id}, status={status}")
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()

    # 使用北京时间记录执行时间
    execution_time = get_beijing_time().isoformat()

    cursor.execute('''
        INSERT INTO task_logs (task_id, status, message, execution_time)
        VALUES (?, ?, ?, ?)
    ''', (task_id, status, message, execution_time))

    conn.commit()
    conn.close()
    logger.info(f"Task execution logged successfully: task_id={task_id}")


def get_condition_text(condition_type, threshold):
    """获取SQL条件文本"""
    conditions = {
        'not_empty': '结果不为空',
        'rows_gt': f'行数大于 {threshold}',
        'rows_lt': f'行数小于 {threshold}',
        'rows_eq': f'行数等于 {threshold}',
        'rows_neq': f'行数不等于 {threshold}'
    }
    return conditions.get(condition_type, condition_type)

def log_sql_alert_execution(alert_id, alert_name, status, message, details=None):
    """记录SQL预警执行日志"""
    try:
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        
        cursor.execute('''
            INSERT INTO sql_alert_logs (alert_id, alert_name, status, message, execution_time, details)
            VALUES (?, ?, ?, ?, ?, ?)
        ''', (alert_id, alert_name, status, message, datetime.now(), details))
        
        conn.commit()
        conn.close()
        logger.info(f"SQL alert execution logged successfully: alert_id={alert_id}, status={status}")
    except Exception as e:
        logger.error(f"Failed to log SQL alert execution: {str(e)}")


def send_task_notification(task_id, status, message):
    """发送任务执行状态预警通知"""
    try:
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        
        # 获取任务信息
        cursor.execute('SELECT name FROM tasks WHERE id = ?', (task_id,))
        task = cursor.fetchone()
        if not task:
            conn.close()
            return
        
        task_name = task[0]
        
        # 映射状态到预警类型
        alert_type = status  # 直接使用状态值，不进行映射
        
        # 获取对应状态的所有预警配置（不再限制特定任务）
        cursor.execute('''
            SELECT ta.recipients, ec.mail_server, ec.mail_port, ec.mail_use_tls, 
                   ec.mail_username, ec.mail_password, ec.mail_default_sender, ec.id
            FROM task_alerts ta
            JOIN email_configs ec ON ta.email_config_id = ec.id
            WHERE ta.alert_type = ? AND ta.is_enabled = 1
        ''', (alert_type,))
        
        alerts = cursor.fetchall()
        
        # 如果没有配置预警，直接返回
        if not alerts:
            conn.close()
            return
        
        # 为每个预警配置发送邮件
        for recipients, mail_server, mail_port, mail_use_tls, mail_username, mail_password, mail_default_sender, email_config_id in alerts:
            try:
                # 构建邮件内容
                status_text = "成功" if status == "success" else "失败"
                subject = f"任务执行{status_text}预警: {task_name}"
                
                body = f"任务执行{status_text}，触发了预警通知。\n\n"
                body += f"任务名称: {task_name}\n"
                body += f"任务ID: {task_id}\n"
                body += f"执行状态: {status_text}\n"
                body += f"执行时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n"
                body += "详细信息:\n"
                body += message
                
                # 发送预警邮件
                with app.app_context():
                    send_alert_email(
                        mail_server, mail_port, mail_use_tls, mail_username, mail_password, mail_default_sender,
                        recipients, subject, body
                    )
                
                # 记录预警日志
                sent_time = get_beijing_time().strftime('%Y-%m-%d %H:%M:%S')
                cursor.execute('''
                    INSERT INTO notification_logs 
                    (task_id, task_name, alert_type, email_config_id, recipients, subject, body, status, sent_time)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (task_id, task_name, status, email_config_id, recipients, subject, body, 'sent', sent_time))
                
                logger.info(f"任务预警邮件已发送 - 任务ID: {task_id}, 状态: {status}, 收件人: {recipients}")
            except Exception as e:
                error_msg = str(e)
                logger.error(f"发送任务预警邮件失败 - 任务ID: {task_id}, 状态: {status}, 错误: {error_msg}")
                
                # 记录预警失败日志
                sent_time = get_beijing_time().strftime('%Y-%m-%d %H:%M:%S')
                cursor.execute('''
                    INSERT INTO notification_logs 
                    (task_id, task_name, alert_type, email_config_id, recipients, subject, body, status, error_message, sent_time)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (task_id, task_name, status, email_config_id, recipients, subject, body, 'failed', error_msg, sent_time))
        
        conn.commit()
        conn.close()
                
    except Exception as e:
        logger.error(f"处理任务预警通知失败 - 任务ID: {task_id}, 状态: {status}, 错误: {str(e)}")


def can_execute_task(cursor, task_id):
    """
    检查任务是否可以执行（所有依赖任务都已完成）
    """
    # 获取任务的依赖关系
    cursor.execute('SELECT dependencies FROM tasks WHERE id = ?', (task_id,))
    task_row = cursor.fetchone()

    if not task_row or not task_row[0]:
        # 没有依赖任务，可以直接执行
        return True

    # 解析依赖任务ID
    try:
        dependency_ids = [int(dep_id) for dep_id in task_row[0].split(',') if dep_id]
    except ValueError:
        # 依赖关系格式错误，允许执行
        logger.warning(f"Task {task_id} has invalid dependency format")
        return True

    if not dependency_ids:
        # 没有有效的依赖任务，可以直接执行
        return True

    # 检查所有依赖任务的最新执行状态
    placeholders = ','.join('?' * len(dependency_ids))
    # 使用子查询获取每个任务的最新日志记录
    cursor.execute(f'''
        SELECT t.id, latest_log.status
        FROM tasks t
        LEFT JOIN (
            SELECT task_id, status
            FROM task_logs tl1
            WHERE tl1.execution_time = (
                SELECT MAX(tl2.execution_time)
                FROM task_logs tl2
                WHERE tl2.task_id = tl1.task_id
            )
        ) latest_log ON t.id = latest_log.task_id
        WHERE t.id IN ({placeholders})
    ''', dependency_ids)

    dependency_results = cursor.fetchall()

    # 检查是否所有依赖任务都成功执行
    for dep_id, status in dependency_results:
        # 如果status为None，说明该任务从未执行过
        if status is None:
            logger.info(f"Task {task_id} cannot execute because dependency task {dep_id} has never been executed")
            return False
        if status != 'success':
            logger.info(f"Task {task_id} cannot execute because dependency task {dep_id} did not succeed")
            return False

    # 所有依赖任务都成功执行
    return True


def task_scheduler():
    """任务调度器"""
    logger.info("Task scheduler started")
    while True:
        try:
            conn = sqlite3.connect(DB_PATH)
            cursor = conn.cursor()

            # 获取所有激活的任务
            now = get_beijing_time()
            cursor.execute('''
                SELECT id, name, task_type, script_path, sql_script_id, schedule_interval, last_run, next_run, max_retries, retry_delay, cron_expression
                FROM tasks 
                WHERE is_active = 1
            ''')

            tasks = cursor.fetchall()

            for task in tasks:
                task_id, name, task_type, script_path, sql_script_id, interval, last_run, next_run, max_retries, retry_delay, cron_expression = task

                # 检查任务是否已经在执行中
                with executing_tasks_lock:
                    if task_id in executing_tasks:
                        logger.info(f"Task {task_id} is already executing, skipping")
                        continue
                    # 将任务标记为正在执行
                    executing_tasks.add(task_id)

                try:
                    # 计算下次运行时间
                    if not next_run:
                        # 如果是第一次运行，设置为现在
                        next_run_time = now
                    else:
                        # 解析存储的下次运行时间
                        next_run_time = datetime.fromisoformat(next_run)

                    # 检查任务是否可以执行（依赖任务已完成）
                    if not can_execute_task(cursor, task_id):
                        continue  # 跳过不能执行的任务

                    # 检查是否需要重试失败的任务
                    should_retry = False
                    if max_retries > 0:
                        # 查找最近一次执行的日志
                        cursor.execute('''
                            SELECT status, execution_time
                            FROM task_logs
                            WHERE task_id = ?
                            ORDER BY execution_time DESC
                            LIMIT 1
                        ''', (task_id,))

                        last_log = cursor.fetchone()

                        # 检查是否需要重试
                        if last_log:
                            last_status, last_execution_time = last_log
                            if last_status == 'failed':
                                # 查找最近连续失败的次数
                                cursor.execute('''
                                    SELECT COUNT(*) as failed_count
                                    FROM (
                                        SELECT status
                                        FROM task_logs
                                        WHERE task_id = ?
                                        ORDER BY execution_time DESC
                                        LIMIT ?
                                    )
                                    WHERE status = 'failed'
                                ''', (task_id, max_retries + 1))

                                failed_count = cursor.fetchone()[0]
                                if failed_count <= max_retries:
                                    # 检查距离上次执行的时间是否超过了重试延迟
                                    last_exec_time = datetime.fromisoformat(last_execution_time)
                                    if (now - last_exec_time).total_seconds() >= retry_delay:
                                        should_retry = True
                                        logger.info(
                                            f"Retrying failed task {task_id} (attempt {failed_count}/{max_retries})")

                    # 如果到了执行时间或者需要重试
                    if now >= next_run_time or should_retry:
                        logger.info(f"Executing scheduled task: {name} (ID: {task_id}, Type: {task_type})")
                        print(f"Executing task: {name} (Type: {task_type})")

                        # 如果不是重试情况，更新任务的执行时间和下次运行时间
                        if not should_retry:
                            new_last_run = now
                            # 如果有cron表达式，使用croniter计算下次运行时间
                            if cron_expression:
                                try:
                                    cron = croniter(cron_expression, now)
                                    new_next_run = cron.get_next(datetime)
                                except Exception as e:
                                    logger.error(
                                        f"Invalid cron expression for task {task_id}: {cron_expression}, error: {e}")
                                    new_next_run = now + timedelta(seconds=interval)  # 回退到使用interval
                            else:
                                # 使用间隔时间计算下次运行时间
                                new_next_run = now + timedelta(seconds=interval)

                            cursor.execute('''
                                UPDATE tasks 
                                SET last_run = ?, next_run = ?
                                WHERE id = ?
                            ''', (new_last_run.isoformat(), new_next_run.isoformat(), task_id))

                            # 提交事务，确保任务状态已更新
                            conn.commit()

                        # 根据任务类型执行相应的操作
                        if task_type == 'python' and script_path:
                            # 执行Python脚本（同步执行，不使用线程）
                            result = execute_script(script_path)
                            # 记录执行日志
                            if result and result.returncode == 0:
                                log_content = f'Python脚本执行成功: {script_path}\n'
                                if result.stdout:
                                    log_content += f'输出:\n{result.stdout}'
                                log_task_execution(task_id, 'success', log_content)
                                # 发送成功预警通知
                                send_task_notification(task_id, 'success', log_content)
                            else:
                                error_msg = result.stderr if result else '未知错误'
                                output_msg = result.stdout if result else ''
                                log_content = f'Python脚本执行失败: {script_path}\n'
                                if error_msg:
                                    log_content += f'错误:\n{error_msg}\n'
                                if output_msg:
                                    log_content += f'输出:\n{output_msg}'
                                log_task_execution(task_id, 'failed', log_content)
                                # 发送失败预警通知
                                send_task_notification(task_id, 'failed', log_content)
                        elif task_type == 'sql' and sql_script_id:
                            # 执行SQL脚本（同步执行，不使用线程）
                            success, result = execute_sql_script(sql_script_id)
                            # 记录执行日志
                            if success:
                                log_content = f'SQL脚本执行成功: {sql_script_id}'
                                log_task_execution(task_id, 'success', log_content)
                                # 发送成功预警通知
                                send_task_notification(task_id, 'success', log_content)
                            else:
                                log_content = f'SQL脚本执行失败: {sql_script_id}, 错误: {result}'
                                log_task_execution(task_id, 'failed', log_content)
                                # 发送失败预警通知
                                send_task_notification(task_id, 'failed', log_content)
                finally:
                    # 无论任务执行成功与否，都要从执行中任务集合中移除
                    with executing_tasks_lock:
                        executing_tasks.discard(task_id)

            conn.close()

            # 检查所有激活的SQL预警
            try:
                conn = sqlite3.connect(DB_PATH)
                cursor = conn.cursor()
                
                # 获取所有激活的SQL预警
                cursor.execute('''
                    SELECT id, name, sql_script_id, email_config_id, recipients, condition_type, threshold, last_check, cron_expression, next_check
                    FROM sql_alerts 
                    WHERE is_enabled = 1
                ''')
                
                sql_alerts = cursor.fetchall()
                
                for alert in sql_alerts:
                    alert_id, name, sql_script_id, email_config_id, recipients, condition_type, threshold, last_check, cron_expression, next_check = alert
                    
                    # 检查SQL预警是否已经在执行中
                    with executing_sql_alerts_lock:
                        if alert_id in executing_sql_alerts:
                            logger.info(f"SQL Alert {alert_id} is already executing, skipping")
                            continue
                        # 将SQL预警标记为正在执行
                        executing_sql_alerts.add(alert_id)
                    
                    try:
                        # 检查是否到了检查时间
                        should_check = False
                        
                        if cron_expression and next_check:
                            # 如果有cron表达式和下次检查时间，检查是否到了时间
                            try:
                                next_check_time = datetime.fromisoformat(next_check)
                                if now >= next_check_time:
                                    should_check = True
                            except:
                                logger.error(f"Invalid next_check time format for SQL alert {alert_id}: {next_check}")
                                should_check = False
                        elif not cron_expression and not next_check:
                            # 如果没有cron表达式和下次检查时间，使用默认间隔（5分钟）
                            if not last_check:
                                should_check = True
                            else:
                                try:
                                    last_check_time = datetime.fromisoformat(last_check)
                                    if (now - last_check_time).total_seconds() >= 300:  # 5分钟
                                        should_check = True
                                except:
                                    logger.error(f"Invalid last_check time format for SQL alert {alert_id}: {last_check}")
                                    should_check = True
                        
                        if should_check:
                            logger.info(f"Checking SQL alert: {name} (ID: {alert_id})")
                            
                            # 执行SQL预警检查
                            result = _check_sql_alert_internal(alert_id)
                            
                            # 记录检查结果
                            if isinstance(result, dict) and 'success' in result and result['success']:
                                if result.get('triggered', False):
                                    if result.get('email_sent', False):
                                        logger.info(f"SQL预警 {alert_id} 触发并发送邮件成功")
                                    else:
                                        logger.warning(f"SQL预警 {alert_id} 触发但邮件发送失败")
                                else:
                                    logger.info(f"SQL预警 {alert_id} 检查完成，未触发预警")
                            else:
                                logger.error(f"SQL预警 {alert_id} 检查失败: {result.get('error', '未知错误')}")
                            
                            # 如果有cron表达式，计算下次检查时间
                            if cron_expression:
                                try:
                                    from croniter import croniter
                                    cron = croniter(cron_expression, now)
                                    new_next_check = cron.get_next(datetime)
                                    
                                    cursor.execute('''
                                        UPDATE sql_alerts 
                                        SET last_check = ?, next_check = ?
                                        WHERE id = ?
                                    ''', (now.isoformat(), new_next_check.isoformat(), alert_id))
                                    
                                    logger.info(f"SQL alert {alert_id} next check time: {new_next_check}")
                                except Exception as e:
                                    logger.error(f"Invalid cron expression for SQL alert {alert_id}: {cron_expression}, error: {e}")
                            else:
                                # 如果没有cron表达式，设置5分钟后的检查时间
                                new_next_check = now + timedelta(seconds=300)
                                
                                cursor.execute('''
                                    UPDATE sql_alerts 
                                    SET last_check = ?, next_check = ?
                                    WHERE id = ?
                                ''', (now.isoformat(), new_next_check.isoformat(), alert_id))
                            
                            conn.commit()
                    finally:
                        # 无论SQL预警检查成功与否，都要从执行中SQL预警集合中移除
                        with executing_sql_alerts_lock:
                            executing_sql_alerts.discard(alert_id)
                
                conn.close()
            except Exception as e:
                logger.error(f"Error checking SQL alerts: {str(e)}")
                try:
                    if 'conn' in locals() and conn:
                        conn.close()
                except:
                    pass

            # 每隔1秒检查一次
            time.sleep(1)

        except Exception as e:
            error_msg = f"Scheduler error: {e}"
            logger.error(error_msg)
            print(error_msg)

            # 记录调度器错误到任务日志
            try:
                conn = sqlite3.connect(DB_PATH)
                cursor = conn.cursor()
                # 获取所有激活的任务
                cursor.execute('''
                    SELECT id FROM tasks 
                    WHERE is_active = 1
                ''')
                tasks = cursor.fetchall()
                conn.close()

                # 为每个任务记录错误日志
                for task in tasks:
                    task_id = task[0]
                    log_task_execution(task_id, 'failed', error_msg)
            except Exception as log_error:
                logger.error(f"Failed to log scheduler error to task logs: {log_error}")

            time.sleep(1)


@app.route('/')
def index():
    # 检查用户是否已登录
    if 'user_id' not in session:
        return redirect(url_for('login'))
    return render_template('index.html')


@app.route('/test-notification-logs')
def test_notification_logs_page():
    # 检查用户是否已登录
    if 'user_id' not in session:
        return redirect(url_for('login'))
    return app.send_static_file('test_notification_logs.html')


@app.route('/test-sql-alerts-table')
def test_sql_alerts_table_page():
    return render_template('test_sql_alerts_table_mock.html')


@app.route('/email-configs')
def email_configs():
    # 检查用户是否已登录
    if 'user_id' not in session:
        return redirect(url_for('login'))
    return render_template('email_configs.html')


# 已删除重复的before_request装饰器


@app.after_request
def log_response_info(response):
    """记录响应信息"""
    logger.info(f"响应: {response.status_code}")
    if response.is_json:
        logger.info(f"响应数据: {response.get_json()}")
    return response


@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.json['username'] if request.is_json else request.form['username']
        password = request.json['password'] if request.is_json else request.form['password']
        
        # 添加调试日志
        logger.info(f"Login attempt: username={username}, is_json={request.is_json}")

        # 验证用户凭据
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        cursor.execute('''
            SELECT id, username, password_hash
            FROM users
            WHERE username = ?
        ''', (username,))

        user = cursor.fetchone()
        conn.close()
        
        # 添加调试日志
        logger.info(f"User found: {user is not None}")
        if user:
            logger.info(f"User ID: {user[0]}, Username: {user[1]}")

        if user and check_password_hash(user[2], password):
            # 登录成功，设置会话
            session['user_id'] = user[0]
            session['username'] = user[1]
            session.permanent = False
            
            # 强制保存会话
            session.modified = True
            
            # 添加调试日志
            logger.info(f"Login successful for user: {username}")
            logger.info(f"Session data: {dict(session)}")

            return jsonify({'message': '登录成功', 'redirect': '/'})
        else:
            # 添加调试日志
            logger.info(f"Login failed for user: {username}")
            
            if request.is_json:
                return jsonify({'error': '用户名或密码错误'}), 401
            else:
                return render_template('login.html', error='用户名或密码错误')

    # GET 请求时显示登录页面
    if 'user_id' in session:
        return redirect(url_for('index'))
    return render_template('login.html')


@app.route('/logout')
def logout():
    # 清除会话数据
    session.clear()
    return redirect(url_for('login'))


# @app.before_request
# def log_request_info():
#     """记录所有请求信息"""
#     logger.info(f"Request: {request.method} {request.path} - Headers: {dict(request.headers)}")
#     if request.method in ['POST', 'PUT', 'PATCH']:
#         logger.info(f"Request data: {request.get_data(as_text=True)}")


# 添加API Token认证装饰器
def api_token_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        # 检查是否有API Token
        token = None
        if 'Authorization' in request.headers:
            auth_header = request.headers['Authorization']
            try:
                # 期望格式: "Bearer TOKEN"
                if auth_header.startswith("Bearer "):
                    token = auth_header.split(" ")[1]
            except (IndexError, ValueError):
                pass

        if not token:
            return jsonify({'error': '缺少API Token，请在Authorization头中使用"Bearer YOUR_TOKEN"格式'}), 401

        # 验证Token
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        cursor.execute('''
            SELECT u.id, u.username 
            FROM user_tokens t
            JOIN users u ON t.user_id = u.id
            WHERE t.token = ? AND t.expires_at > datetime('now')
        ''', (token,))

        user = cursor.fetchone()
        conn.close()

        if not user:
            return jsonify({'error': '无效或过期的API Token'}), 401

        # 将用户信息添加到请求上下文中
        request.current_user_id = user[0]
        request.current_username = user[1]

        return f(*args, **kwargs)

    return decorated_function


# 为所有API路由添加登录检查装饰器

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return jsonify({'error': '未授权访问'}), 401
        return f(*args, **kwargs)

    return decorated_function


@app.route('/email-config')
@login_required
def email_config():
    return render_template('email_config.html')


@app.route('/email-test')
@login_required
def email_test():
    return render_template('email_test.html')


@app.route('/alerts')
@login_required
def alerts_page():
    """预警配置页面"""
    return render_template('alerts.html')


@app.route('/alerts-debug')
def alerts_debug():
    """预警配置调试页面"""
    return render_template('alerts_debug.html')


@app.route('/db-config')
@login_required
def db_config_page():
    """数据库配置页面"""
    return render_template('db_config.html')


@app.route('/sql-scripts')
@login_required
def sql_scripts_page():
    """SQL脚本页面"""
    return render_template('sql_scripts.html')


@app.route('/users')
@login_required
def users_page():
    """用户管理页面"""
    return render_template('users.html')


@app.route('/notification-logs')
@login_required
def notification_logs_page():
    """预警日志页面"""
    return render_template('notification_logs.html')


@app.route('/api/email-config', methods=['GET'])
@login_required
def get_email_config():
    """获取邮件配置"""
    try:
        # 连接数据库
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        
        # 获取所有邮件配置
        cursor.execute('''
            SELECT id, config_name, mail_server, mail_port, mail_use_tls, 
                   mail_username, mail_default_sender, mail_notification_enabled, is_default
            FROM email_configs
            ORDER BY is_default DESC, created_at ASC
        ''')
        configs = cursor.fetchall()
        
        # 转换为字典列表
        config_list = []
        for config in configs:
            config_dict = {
                'id': config[0],
                'config_name': config[1],
                'mail_server': config[2],
                'mail_port': config[3],
                'mail_use_tls': bool(config[4]),
                'mail_username': config[5],
                'mail_default_sender': config[6],
                'mail_notification_enabled': bool(config[7]),
                'is_default': bool(config[8])
            }
            config_list.append(config_dict)
        
        # 如果有配置，获取默认配置
        default_config = None
        if config_list:
            for config in config_list:
                if config['is_default']:
                    default_config = config
                    break
            
            # 如果没有默认配置，使用第一个配置
            if not default_config:
                default_config = config_list[0]
        
        conn.close()
        
        # 返回配置列表和默认配置
        return jsonify({
            'configs': config_list,
            'default_config': default_config
        })
    except Exception as e:
        logger.error(f"获取邮件配置失败: {e}")
        return jsonify({'error': f'获取邮件配置失败: {str(e)}'}), 500


@app.route('/api/email-configs', methods=['GET'])
@login_required
def get_email_configs():
    """获取邮件配置列表"""
    try:
        # 连接数据库
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        
        # 获取所有邮件配置
        cursor.execute('''
            SELECT id, config_name, mail_server, mail_port, mail_use_tls, 
                   mail_username, mail_default_sender, mail_notification_enabled, is_default, created_at
            FROM email_configs
            ORDER BY is_default DESC, created_at ASC
        ''')
        configs = cursor.fetchall()
        
        # 转换为字典列表
        config_list = []
        for config in configs:
            config_dict = {
                'id': config[0],
                'config_name': config[1],
                'mail_server': config[2],
                'mail_port': config[3],
                'mail_use_tls': bool(config[4]),
                'mail_username': config[5],
                'mail_default_sender': config[6],
                'mail_notification_enabled': bool(config[7]),
                'is_default': bool(config[8]),
                'created_at': config[9]
            }
            config_list.append(config_dict)
        
        conn.close()
        
        # 返回配置列表
        return jsonify(config_list)
    except Exception as e:
        logger.error(f"获取邮件配置列表失败: {e}")
        return jsonify({'error': f'获取邮件配置列表失败: {str(e)}'}), 500


@app.route('/api/email-config/<int:config_id>', methods=['GET'])
@login_required
def get_email_config_by_id(config_id):
    """根据ID获取单个邮件配置"""
    try:
        # 连接数据库
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        
        # 获取指定ID的邮件配置
        cursor.execute('''
            SELECT id, config_name, mail_server, mail_port, mail_use_tls, 
                   mail_username, mail_password, mail_default_sender, mail_notification_enabled, is_default
            FROM email_configs
            WHERE id = ?
        ''', (config_id,))
        config = cursor.fetchone()
        
        if not config:
            conn.close()
            return jsonify({'error': '邮件配置不存在'}), 404
        
        # 转换为字典
        config_dict = {
            'id': config[0],
            'config_name': config[1],
            'mail_server': config[2],
            'mail_port': config[3],
            'mail_use_tls': bool(config[4]),
            'mail_username': config[5],
            'mail_password': config[6],
            'mail_default_sender': config[7],
            'mail_notification_enabled': bool(config[8]),
            'is_default': bool(config[9])
        }
        
        conn.close()
        
        return jsonify(config_dict)
    except Exception as e:
        logger.error(f"获取邮件配置失败: {e}")
        return jsonify({'error': f'获取邮件配置失败: {str(e)}'}), 500


@app.route('/api/email-config/<int:config_id>', methods=['DELETE'])
@login_required
def delete_email_config(config_id):
    """删除邮件配置"""
    try:
        # 连接数据库
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        
        # 检查是否是默认配置
        cursor.execute("SELECT is_default FROM email_configs WHERE id = ?", (config_id,))
        config = cursor.fetchone()
        
        if not config:
            conn.close()
            return jsonify({'error': '邮件配置不存在'}), 404
        
        # 如果是默认配置，不允许删除
        if config[0]:
            conn.close()
            return jsonify({'error': '不能删除默认邮件配置'}), 400
        
        # 删除配置
        cursor.execute("DELETE FROM email_configs WHERE id = ?", (config_id,))
        conn.commit()
        conn.close()
        
        logger.info(f"邮件配置已删除: ID={config_id}")
        return jsonify({'message': '邮件配置删除成功'})
    except Exception as e:
        logger.error(f"删除邮件配置失败: {e}")
        return jsonify({'error': f'删除邮件配置失败: {str(e)}'}), 500


@app.route('/api/email-config/<int:config_id>/set-default', methods=['POST'])
@login_required
def set_default_email_config(config_id):
    """设置默认邮件配置"""
    try:
        # 连接数据库
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        
        # 检查配置是否存在
        cursor.execute("SELECT id FROM email_configs WHERE id = ?", (config_id,))
        config = cursor.fetchone()
        
        if not config:
            conn.close()
            return jsonify({'error': '邮件配置不存在'}), 404
        
        # 将所有配置的默认标志设为False
        cursor.execute("UPDATE email_configs SET is_default = 0")
        
        # 将指定配置设为默认
        cursor.execute("UPDATE email_configs SET is_default = 1 WHERE id = ?", (config_id,))
        conn.commit()
        conn.close()
        
        # 更新Flask应用的邮件配置
        cursor = sqlite3.connect(DB_PATH).cursor()
        cursor.execute('''
            SELECT mail_server, mail_port, mail_use_tls, mail_username, 
                   mail_password, mail_default_sender, mail_notification_enabled
            FROM email_configs
            WHERE id = ?
        ''', (config_id,))
        config_data = cursor.fetchone()
        cursor.connection.close()
        
        if config_data:
            app.config['MAIL_SERVER'] = config_data[0]
            app.config['MAIL_PORT'] = config_data[1]
            app.config['MAIL_USE_TLS'] = bool(config_data[2])
            app.config['MAIL_USERNAME'] = config_data[3]
            app.config['MAIL_PASSWORD'] = config_data[4]
            app.config['MAIL_DEFAULT_SENDER'] = config_data[5]
            app.config['MAIL_NOTIFICATION_ENABLED'] = bool(config_data[6])
            
            # 重新初始化邮件对象
            global mail
            mail.init_app(app)
        
        logger.info(f"默认邮件配置已设置: ID={config_id}")
        return jsonify({'message': '默认邮件配置设置成功'})
    except Exception as e:
        logger.error(f"设置默认邮件配置失败: {e}")
        return jsonify({'error': f'设置默认邮件配置失败: {str(e)}'}), 500


@app.route('/api/email-config', methods=['POST'])
@login_required
def save_email_config():
    """保存邮件配置"""
    try:
        data = request.get_json()
        config_id = data.get('id')  # 如果有ID，则为更新操作
        
        # 连接数据库
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        
        # 如果设置为默认配置，先将其他配置的默认标志设为False
        if data.get('is_default', False):
            cursor.execute("UPDATE email_configs SET is_default = 0")
        
        if config_id:
            # 更新现有配置
            cursor.execute('''
                UPDATE email_configs 
                SET config_name = ?, mail_server = ?, mail_port = ?, mail_use_tls = ?, 
                    mail_username = ?, mail_password = ?, mail_default_sender = ?, 
                    mail_notification_enabled = ?, is_default = ?, updated_at = CURRENT_TIMESTAMP
                WHERE id = ?
            ''', (
                data.get('config_name', ''),
                data.get('mail_server', ''),
                int(data.get('mail_port', 587)),
                data.get('mail_use_tls', True),
                data.get('mail_username', ''),
                data.get('mail_password', ''),
                data.get('mail_default_sender', ''),
                data.get('mail_notification_enabled', False),
                data.get('is_default', False),
                config_id
            ))
        else:
            # 插入新配置
            cursor.execute('''
                INSERT INTO email_configs 
                (config_name, mail_server, mail_port, mail_use_tls, mail_username, mail_password, 
                 mail_default_sender, mail_notification_enabled, is_default)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                data.get('config_name', ''),
                data.get('mail_server', ''),
                int(data.get('mail_port', 587)),
                data.get('mail_use_tls', True),
                data.get('mail_username', ''),
                data.get('mail_password', ''),
                data.get('mail_default_sender', ''),
                data.get('mail_notification_enabled', False),
                data.get('is_default', False)
            ))
        
        conn.commit()
        conn.close()
        
        # 更新Flask应用的邮件配置
        app.config['MAIL_SERVER'] = data.get('mail_server', '')
        app.config['MAIL_PORT'] = int(data.get('mail_port', 587))
        app.config['MAIL_USE_TLS'] = data.get('mail_use_tls', True)
        app.config['MAIL_USERNAME'] = data.get('mail_username', '')
        app.config['MAIL_PASSWORD'] = data.get('mail_password', '')
        app.config['MAIL_DEFAULT_SENDER'] = data.get('mail_default_sender', '')
        app.config['MAIL_NOTIFICATION_ENABLED'] = data.get('mail_notification_enabled', False)
        
        # 重新初始化邮件对象
        global mail
        mail.init_app(app)
        
        logger.info(f"邮件配置已保存到数据库: {data.get('config_name', '')}")
        return jsonify({'message': '邮件配置保存成功'})
    except Exception as e:
        logger.error(f"保存邮件配置失败: {e}")
        return jsonify({'error': f'保存邮件配置失败: {str(e)}'}), 500


@app.route('/api/email-configs', methods=['POST'])
@login_required
def create_email_config():
    """创建新的邮件配置"""
    try:
        data = request.get_json()
        
        # 连接数据库
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        
        # 如果设为默认，先将其他配置设为非默认
        if data.get('is_default'):
            cursor.execute("UPDATE email_configs SET is_default = 0")
        
        # 插入新配置
        cursor.execute('''
            INSERT INTO email_configs 
            (config_name, mail_server, mail_port, mail_use_tls, mail_username, mail_password, 
             mail_default_sender, mail_notification_enabled, is_default, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, datetime('now'), datetime('now'))
        ''', (
            data.get('config_name'),
            data.get('mail_server'),
            data.get('mail_port'),
            data.get('mail_use_tls'),
            data.get('mail_username'),
            data.get('mail_password'),
            data.get('mail_default_sender'),
            data.get('mail_notification_enabled', True),
            data.get('is_default', 0)
        ))
        
        conn.commit()
        conn.close()
        
        logger.info(f"邮件配置已创建: {data.get('name', '')}")
        return jsonify({'message': '邮件配置创建成功'})
    except Exception as e:
        logger.error(f"创建邮件配置失败: {e}")
        return jsonify({'error': f'创建邮件配置失败: {str(e)}'}), 500


@app.route('/api/email-configs/<int:config_id>', methods=['GET'])
@login_required
def get_email_config_by_id_plural(config_id):
    """根据ID获取单个邮件配置（复数形式）"""
    try:
        # 连接数据库
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        
        # 获取指定ID的邮件配置
        cursor.execute('''
            SELECT id, config_name, mail_server, mail_port, mail_use_tls, 
                   mail_username, mail_password, mail_default_sender, mail_notification_enabled, is_default
            FROM email_configs
            WHERE id = ?
        ''', (config_id,))
        config = cursor.fetchone()
        
        conn.close()
        
        if not config:
            return jsonify({'error': '找不到指定的邮件配置'}), 404
        
        # 转换为字典
        config_dict = {
            'id': config[0],
            'config_name': config[1],
            'mail_server': config[2],
            'mail_port': config[3],
            'mail_use_tls': bool(config[4]),
            'mail_username': config[5],
            'mail_password': config[6],
            'mail_default_sender': config[7],
            'mail_notification_enabled': bool(config[8]),
            'is_default': bool(config[9])
        }
        
        return jsonify(config_dict)
    except Exception as e:
        logger.error(f"获取邮件配置失败: {e}")
        return jsonify({'error': f'获取邮件配置失败: {str(e)}'}), 500


@app.route('/api/email-configs/<int:config_id>', methods=['PUT'])
@login_required
def update_email_config_plural(config_id):
    """更新邮件配置（复数形式）"""
    try:
        data = request.get_json()
        
        # 连接数据库
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        
        # 构建更新语句
        update_fields = []
        update_values = []
        
        if data.get('config_name'):
            update_fields.append('config_name = ?')
            update_values.append(data.get('config_name'))
        
        if data.get('mail_server'):
            update_fields.append('mail_server = ?')
            update_values.append(data.get('mail_server'))
        
        if data.get('mail_port'):
            update_fields.append('mail_port = ?')
            update_values.append(data.get('mail_port'))
        
        if data.get('mail_use_tls') is not None:
            update_fields.append('mail_use_tls = ?')
            update_values.append(data.get('mail_use_tls'))
        
        if data.get('mail_username'):
            update_fields.append('mail_username = ?')
            update_values.append(data.get('mail_username'))
        
        if data.get('mail_password'):
            update_fields.append('mail_password = ?')
            update_values.append(data.get('mail_password'))
        
        if data.get('mail_default_sender'):
            update_fields.append('mail_default_sender = ?')
            update_values.append(data.get('mail_default_sender'))
        
        if data.get('mail_notification_enabled') is not None:
            update_fields.append('mail_notification_enabled = ?')
            update_values.append(data.get('mail_notification_enabled'))
        
        # 如果设为默认，先将其他配置设为非默认
        if data.get('is_default'):
            cursor.execute("UPDATE email_configs SET is_default = 0")
            update_fields.append('is_default = ?')
            update_values.append(1)
        elif data.get('is_default') is not None:
            update_fields.append('is_default = ?')
            update_values.append(0)
        
        # 添加更新时间
        update_fields.append('updated_at = datetime(\'now\')')
        
        # 添加配置ID到更新值列表
        update_values.append(config_id)
        
        # 执行更新
        cursor.execute(f'''
            UPDATE email_configs 
            SET {', '.join(update_fields)}
            WHERE id = ?
        ''', update_values)
        
        conn.commit()
        conn.close()
        
        logger.info(f"邮件配置已更新: {config_id}")
        return jsonify({'message': '邮件配置更新成功'})
    except Exception as e:
        logger.error(f"更新邮件配置失败: {e}")
        return jsonify({'error': f'更新邮件配置失败: {str(e)}'}), 500


@app.route('/api/email-configs/<int:config_id>', methods=['DELETE'])
@login_required
def delete_email_config_plural(config_id):
    """删除邮件配置（复数形式）"""
    try:
        # 连接数据库
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        
        # 检查是否为默认配置
        cursor.execute("SELECT is_default FROM email_configs WHERE id = ?", (config_id,))
        config = cursor.fetchone()
        
        if not config:
            conn.close()
            return jsonify({'error': '找不到指定的邮件配置'}), 404
        
        is_default = config[0]
        
        # 如果是默认配置，不允许删除
        if is_default:
            conn.close()
            return jsonify({'error': '不能删除默认邮件配置，请先设置其他配置为默认'}), 400
        
        # 删除配置
        cursor.execute("DELETE FROM email_configs WHERE id = ?", (config_id,))
        
        conn.commit()
        conn.close()
        
        logger.info(f"邮件配置已删除: {config_id}")
        return jsonify({'message': '邮件配置删除成功'})
    except Exception as e:
        logger.error(f"删除邮件配置失败: {e}")
        return jsonify({'error': f'删除邮件配置失败: {str(e)}'}), 500


@app.route('/api/email-configs/<int:config_id>/test', methods=['POST'])
@login_required
def test_email_config_plural(config_id):
    """测试邮件配置（复数形式）"""
    try:
        data = request.get_json()
        recipient = data.get('recipient')
        
        if not recipient:
            return jsonify({'error': '请提供收件人邮箱'}), 400
        
        # 连接数据库获取配置
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        cursor.execute('''
            SELECT mail_server, mail_port, mail_use_tls, mail_username, mail_password, mail_default_sender
            FROM email_configs
            WHERE id = ?
        ''', (config_id,))
        config = cursor.fetchone()
        conn.close()
        
        if not config:
            return jsonify({'error': '找不到指定的邮件配置'}), 404
            
        mail_server, mail_port, mail_use_tls, mail_username, mail_password, mail_default_sender = config
        mail_use_tls = bool(mail_use_tls)
        
        if not all([mail_server, mail_username, mail_password]):
            return jsonify({'error': '邮件配置不完整'}), 400
        
        # 创建邮件对象
        test_mail = Mail(app)
        app.config['MAIL_SERVER'] = mail_server
        app.config['MAIL_PORT'] = mail_port
        app.config['MAIL_USE_TLS'] = mail_use_tls
        app.config['MAIL_USERNAME'] = mail_username
        app.config['MAIL_PASSWORD'] = mail_password
        app.config['MAIL_DEFAULT_SENDER'] = mail_default_sender or mail_username
        test_mail.init_app(app)
        
        # 发送测试邮件
        msg = Message(
            subject="测试邮件",
            sender=app.config['MAIL_DEFAULT_SENDER'],
            recipients=[recipient]
        )
        msg.body = "这是一封测试邮件，用于验证邮件配置是否正确。"
        
        test_mail.send(msg)
        
        logger.info(f"测试邮件已发送至: {recipient}")
        return jsonify({'message': '测试邮件发送成功'})
    except Exception as e:
        logger.error(f"测试邮件发送失败: {e}")
        return jsonify({'error': f'测试邮件发送失败: {str(e)}'}), 500


@app.route('/api/test-email', methods=['POST'])
@login_required
def test_email():
    """测试邮件发送"""
    try:
        data = request.get_json()
        recipient = data.get('recipient')
        subject = data.get('subject', "测试邮件")
        body = data.get('body', "这是一封测试邮件，用于验证邮件配置是否正确。")
        config_id = data.get('config_id')
        
        if not recipient:
            return jsonify({'error': '请提供收件人邮箱'}), 400
        
        # 如果提供了配置ID，则从数据库获取该配置
        if config_id:
            conn = sqlite3.connect(DB_PATH)
            cursor = conn.cursor()
            cursor.execute('''
                SELECT mail_server, mail_port, mail_use_tls, mail_username, mail_password, mail_default_sender
                FROM email_configs
                WHERE id = ?
            ''', (config_id,))
            config = cursor.fetchone()
            conn.close()
            
            if not config:
                return jsonify({'error': '找不到指定的邮件配置'}), 404
                
            mail_server, mail_port, mail_use_tls, mail_username, mail_password, mail_default_sender = config
            mail_use_tls = bool(mail_use_tls)
        else:
            # 使用全局邮件配置
            mail_server = app.config.get('MAIL_SERVER')
            mail_port = app.config.get('MAIL_PORT')
            mail_username = app.config.get('MAIL_USERNAME')
            mail_password = app.config.get('MAIL_PASSWORD')
            mail_use_tls = app.config.get('MAIL_USE_TLS', True)
            mail_default_sender = app.config.get('MAIL_DEFAULT_SENDER') or mail_username
        
        if not all([mail_server, mail_username, mail_password]):
            return jsonify({'error': '邮件配置不完整，请先配置邮件服务器'}), 400
        
        # 记录配置信息（隐藏密码）
        logger.info(f"尝试连接邮件服务器: {mail_server}:{mail_port}, 用户名: {mail_username}, TLS: {mail_use_tls}")
        
        # 创建测试邮件
        # 重新初始化邮件对象以确保使用最新配置
        global mail
        app.config['MAIL_SERVER'] = mail_server
        app.config['MAIL_PORT'] = mail_port
        app.config['MAIL_USERNAME'] = mail_username
        app.config['MAIL_PASSWORD'] = mail_password
        app.config['MAIL_USE_TLS'] = mail_use_tls
        app.config['MAIL_DEFAULT_SENDER'] = mail_default_sender or mail_username
        mail.init_app(app)
        
        msg = Message(
            subject=subject,
            recipients=[recipient],
            body=body,
            sender=mail_default_sender or mail_username
        )
        
        # 发送邮件
        mail.send(msg)
        
        logger.info(f"测试邮件已发送至: {recipient}")
        return jsonify({
            'success': True,
            'message': f'测试邮件已成功发送至 {recipient}',
            'details': {
                'recipient': recipient,
                'subject': subject,
                'server': f"{mail_server}:{mail_port}",
                'username': mail_username,
                'tls_enabled': mail_use_tls
            }
        })
    except smtplib.SMTPAuthenticationError as e:
        error_msg = f"邮件认证失败，请检查用户名和密码。如果是Gmail，请使用应用专用密码。"
        logger.error(f"{error_msg} 错误详情: {str(e)}")
        return jsonify({
            'success': False,
            'error': error_msg,
            'error_type': 'authentication',
            'details': {
                'server': f"{mail_server}:{mail_port}",
                'username': mail_username,
                'error_code': str(e)
            }
        }), 500

# 任务状态预警API
@app.route('/api/task-alerts', methods=['GET'])
@login_required
def get_task_alerts():
    """获取任务状态预警列表"""
    try:
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        
        cursor.execute('''
            SELECT ta.*, t.name as task_name, ec.config_name as email_config_name
            FROM task_alerts ta
            LEFT JOIN tasks t ON ta.task_id = t.id
            JOIN email_configs ec ON ta.email_config_id = ec.id
            ORDER BY ta.created_at DESC
        ''')
        
        alerts = [dict(row) for row in cursor.fetchall()]
        
        # 添加前端期望的字段
        for alert in alerts:
            # 如果数据库中没有name字段，则使用默认值
            if not alert.get('name'):
                alert['name'] = f"任务预警-{alert['alert_type']}"
            alert['condition'] = alert['alert_type']
            alert['emails'] = alert['recipients']
            alert['enabled'] = alert['is_enabled']
            # 如果没有task_name，使用默认值
            if not alert['task_name']:
                alert['task_name'] = '未指定任务'
        
        conn.close()
        return jsonify({'alerts': alerts})
    except Exception as e:
        logger.error(f"获取任务预警列表失败: {str(e)}")
        return jsonify({'error': '获取任务预警列表失败'}), 500

@app.route('/api/task-alerts', methods=['POST'])
@login_required
def create_task_alert():
    """创建任务状态预警"""
    try:
        data = request.get_json()
        
        # 支持前端发送的字段名
        name = data.get('name')
        task_id = data.get('task_id')
        alert_type = data.get('alert_type') or data.get('condition')  # 'success' 或 'failure'
        email_config_id = data.get('email_config_id')
        recipients = data.get('recipients') or data.get('emails')  # 逗号分隔的邮箱列表
        is_enabled = data.get('is_enabled') or data.get('enabled', True)
        
        # 如果没有提供email_config_id，使用默认的邮件配置
        if not email_config_id:
            conn = sqlite3.connect(DB_PATH)
            cursor = conn.cursor()
            cursor.execute("SELECT id FROM email_configs WHERE is_default = 1 LIMIT 1")
            default_config = cursor.fetchone()
            if default_config:
                email_config_id = default_config[0]
            else:
                # 如果没有默认配置，使用第一个配置
                cursor.execute("SELECT id FROM email_configs ORDER BY id ASC LIMIT 1")
                first_config = cursor.fetchone()
                if first_config:
                    email_config_id = first_config[0]
                else:
                    return jsonify({'error': '请先配置邮件服务器'}), 400
            conn.close()
        
        if not all([alert_type, email_config_id, recipients]):
            return jsonify({'error': '缺少必要参数'}), 400
        
        if alert_type not in ['success', 'failure', 'timeout']:
            return jsonify({'error': '预警类型必须是 success、failure 或 timeout'}), 400
        
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        
        cursor.execute('''
            INSERT INTO task_alerts (name, task_id, alert_type, email_config_id, recipients, is_enabled)
            VALUES (?, ?, ?, ?, ?, ?)
        ''', (name, task_id, alert_type, email_config_id, recipients, is_enabled))
        
        alert_id = cursor.lastrowid
        conn.commit()
        conn.close()
        
        return jsonify({
            'id': alert_id,
            'message': '任务预警创建成功'
        }), 201
    except Exception as e:
        logger.error(f"创建任务预警失败: {str(e)}")
        return jsonify({'error': f'创建任务预警失败: {str(e)}'}), 500

@app.route('/api/task-alerts/<int:alert_id>', methods=['GET'])
@login_required
def get_task_alert_by_id(alert_id):
    """根据ID获取单个任务状态预警"""
    try:
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        
        cursor.execute('''
            SELECT ta.*, ec.config_name as email_config_name
            FROM task_alerts ta
            JOIN email_configs ec ON ta.email_config_id = ec.id
            WHERE ta.id = ?
        ''', (alert_id,))
        
        alert = cursor.fetchone()
        
        conn.close()
        
        if not alert:
            return jsonify({'error': '找不到指定的任务预警'}), 404
        
        # 转换为字典并添加兼容字段
        alert_dict = dict(alert)
        # 添加前端期望的字段
        # 如果数据库中没有name字段，则使用默认值
        if not alert_dict.get('name'):
            alert_dict['name'] = f"任务预警-{alert_dict['alert_type']}"
        alert_dict['condition'] = alert_dict['alert_type']
        alert_dict['emails'] = alert_dict['recipients']
        alert_dict['enabled'] = alert_dict['is_enabled']
        
        return jsonify(alert_dict)
    except Exception as e:
        logger.error(f"获取任务预警失败: {e}")
        return jsonify({'error': '获取任务预警失败'}), 500

@app.route('/api/task-alerts/<int:alert_id>', methods=['PUT'])
@login_required
def update_task_alert(alert_id):
    """更新任务状态预警"""
    try:
        data = request.get_json()
        
        # 支持前端发送的字段名
        name = data.get('name')
        alert_type = data.get('alert_type') or data.get('condition')
        task_id = data.get('task_id')
        email_config_id = data.get('email_config_id')
        recipients = data.get('recipients') or data.get('emails')
        is_enabled = data.get('is_enabled') or data.get('enabled')
        
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        
        # 构建更新语句
        update_fields = []
        update_values = []
        
        if name is not None:
            update_fields.append('name = ?')
            update_values.append(name)
        
        if alert_type is not None:
            update_fields.append('alert_type = ?')
            update_values.append(alert_type)
        
        if task_id is not None:
            update_fields.append('task_id = ?')
            update_values.append(task_id)
        
        if email_config_id is not None:
            update_fields.append('email_config_id = ?')
            update_values.append(email_config_id)
        
        if recipients is not None:
            update_fields.append('recipients = ?')
            update_values.append(recipients)
        
        if is_enabled is not None:
            # 检查当前预警状态
            cursor.execute('SELECT is_enabled FROM task_alerts WHERE id = ?', (alert_id,))
            current_status = cursor.fetchone()
            logger.info(f"任务预警 {alert_id} 当前状态: {current_status[0] if current_status else None}, 请求状态: {is_enabled}")
            
            # 只有当状态不同时才更新
            if current_status and current_status[0] != is_enabled:
                update_fields.append('is_enabled = ?')
                update_values.append(is_enabled)
                logger.info(f"任务预警 {alert_id} 状态将从 {current_status[0]} 更新为 {is_enabled}")
            else:
                logger.info(f"任务预警 {alert_id} 状态未改变，不需要更新")
        
        if not update_fields:
            logger.info(f"任务预警 {alert_id} 没有字段需要更新")
            conn.close()
            # 如果没有字段需要更新，返回成功响应
            return jsonify({'success': True, 'message': '状态未改变'})
        
        update_fields.append('updated_at = CURRENT_TIMESTAMP')
        update_values.append(alert_id)
        
        cursor.execute(f'''
            UPDATE task_alerts
            SET {', '.join(update_fields)}
            WHERE id = ?
        ''', update_values)
        
        conn.commit()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': '任务预警更新成功'
        })
    except Exception as e:
        logger.error(f"更新任务预警失败: {str(e)}")
        return jsonify({'error': '更新任务预警失败'}), 500

@app.route('/api/task-alerts/<int:alert_id>', methods=['DELETE'])
@login_required
def delete_task_alert(alert_id):
    """删除任务状态预警"""
    try:
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        
        cursor.execute('DELETE FROM task_alerts WHERE id = ?', (alert_id,))
        
        if cursor.rowcount == 0:
            conn.close()
            return jsonify({'error': '预警不存在'}), 404
        
        conn.commit()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': '任务预警删除成功'
        })
    except Exception as e:
        logger.error(f"删除任务预警失败: {str(e)}")
        return jsonify({'error': '删除任务预警失败'}), 500

# SQL查询预警API
@app.route('/api/sql-alerts', methods=['GET'])
@login_required
def get_sql_alerts():
    """获取SQL查询预警列表"""
    try:
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        
        cursor.execute('''
            SELECT sa.*, ss.name as sql_script_name, ec.config_name as email_config_name
            FROM sql_alerts sa
            JOIN sql_scripts ss ON sa.sql_script_id = ss.id
            JOIN email_configs ec ON sa.email_config_id = ec.id
            ORDER BY sa.created_at DESC
        ''')
        
        alerts = [dict(row) for row in cursor.fetchall()]
        
        # 添加前端期望的字段
        for alert in alerts:
            alert['emails'] = alert['recipients']
            alert['enabled'] = bool(alert['is_enabled'])  # 转换为布尔值
            alert['script_name'] = alert['sql_script_name']
            # 添加condition和threshold字段，用于前端显示
            alert['condition'] = alert.get('condition_type', 'not_empty')
            alert['threshold'] = alert.get('threshold', 1)
            # 添加cron_expression和next_check字段
            alert['cron_expression'] = alert.get('cron_expression')
            alert['next_check'] = alert.get('next_check')
        
        conn.close()
        return jsonify({'alerts': alerts})
    except Exception as e:
        logger.error(f"获取SQL预警列表失败: {str(e)}")
        return jsonify({'error': '获取SQL预警列表失败'}), 500


@app.route('/api/sql-alerts-debug', methods=['GET'])
def get_sql_alerts_debug():
    """获取SQL查询预警列表 - 调试版本（不需要认证）"""
    try:
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        
        cursor.execute('''
            SELECT sa.*, ss.name as sql_script_name, ec.config_name as email_config_name
            FROM sql_alerts sa
            JOIN sql_scripts ss ON sa.sql_script_id = ss.id
            JOIN email_configs ec ON sa.email_config_id = ec.id
            ORDER BY sa.created_at DESC
        ''')
        
        alerts = [dict(row) for row in cursor.fetchall()]
        
        # 添加前端期望的字段
        for alert in alerts:
            alert['emails'] = alert['recipients']
            alert['enabled'] = bool(alert['is_enabled'])  # 转换为布尔值
            alert['script_name'] = alert['sql_script_name']
            # 添加condition和threshold字段，用于前端显示
            alert['condition'] = alert.get('condition_type', 'not_empty')
            alert['threshold'] = alert.get('threshold', 1)
            # 添加cron_expression和next_check字段
            alert['cron_expression'] = alert.get('cron_expression')
            alert['next_check'] = alert.get('next_check')
        
        conn.close()
        return jsonify({'alerts': alerts})
    except Exception as e:
        logger.error(f"获取SQL预警列表失败: {str(e)}")
        return jsonify({'error': '获取SQL预警列表失败'}), 500

@app.route('/api/sql-alerts', methods=['POST'])
@login_required
def create_sql_alert():
    """创建SQL查询预警"""
    try:
        data = request.get_json()
        
        # 支持前端发送的字段名
        name = data.get('name')
        sql_script_id = data.get('sql_script_id') or data.get('script_id')
        email_config_id = data.get('email_config_id')
        recipients = data.get('recipients') or data.get('emails')  # 逗号分隔的邮箱列表
        condition_type = data.get('condition_type') or data.get('condition', 'not_empty')
        threshold = data.get('threshold', 1)
        is_enabled = data.get('is_enabled') or data.get('enabled', True)
        cron_expression = data.get('cron_expression')
        
        if not all([name, sql_script_id, recipients]):
            return jsonify({'error': '缺少必要参数'}), 400
        
        # 检查SQL脚本是否有关联的数据库配置
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        cursor.execute('SELECT db_config_id FROM sql_scripts WHERE id = ?', (sql_script_id,))
        script_result = cursor.fetchone()
        if not script_result:
            conn.close()
            return jsonify({'error': 'SQL脚本不存在'}), 400
            
        db_config_id = script_result[0]
        if not db_config_id:
            conn.close()
            return jsonify({'error': 'SQL脚本没有关联的数据库配置，请先编辑SQL脚本并选择数据库配置'}), 400
        conn.close()
        
        # 如果没有提供email_config_id，使用默认的邮件配置
        if not email_config_id:
            conn = sqlite3.connect(DB_PATH)
            cursor = conn.cursor()
            cursor.execute("SELECT id FROM email_configs WHERE is_default = 1 LIMIT 1")
            default_config = cursor.fetchone()
            if default_config:
                email_config_id = default_config[0]
            else:
                # 如果没有默认配置，使用第一个配置
                cursor.execute("SELECT id FROM email_configs ORDER BY id ASC LIMIT 1")
                first_config = cursor.fetchone()
                if first_config:
                    email_config_id = first_config[0]
                else:
                    return jsonify({'error': '请先配置邮件服务器'}), 400
            conn.close()
        
        # 计算下次检查时间（如果提供了cron表达式）
        next_check = None
        if cron_expression:
            try:
                from croniter import croniter
                cron = croniter(cron_expression, datetime.now())
                next_check = cron.get_next(datetime)
                logger.info(f"SQL预警 {name} 下次检查时间: {next_check}")
            except Exception as e:
                logger.error(f"Invalid cron expression for SQL alert {name}: {cron_expression}, error: {e}")
                return jsonify({'error': f'无效的cron表达式: {cron_expression}'}), 400
        
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        
        cursor.execute('''
            INSERT INTO sql_alerts (name, sql_script_id, email_config_id, recipients, condition_type, threshold, is_enabled, cron_expression, next_check)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (name, sql_script_id, email_config_id, recipients, condition_type, threshold, is_enabled, cron_expression, next_check))
        
        alert_id = cursor.lastrowid
        conn.commit()
        conn.close()
        
        return jsonify({
            'id': alert_id,
            'message': 'SQL预警创建成功'
        }), 201
    except Exception as e:
        logger.error(f"创建SQL预警失败: {str(e)}")
        return jsonify({'error': f'创建SQL预警失败: {str(e)}'}), 500
    except Exception as e:
        logger.error(f"创建SQL预警失败: {str(e)}")
        return jsonify({'error': '创建SQL预警失败'}), 500

@app.route('/api/sql-alerts/<int:alert_id>', methods=['GET'])
@login_required
def get_sql_alert_by_id(alert_id):
    """根据ID获取单个SQL查询预警"""
    try:
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        
        cursor.execute('''
            SELECT sa.*, ss.name as sql_script_name, ec.config_name as email_config_name
            FROM sql_alerts sa
            JOIN sql_scripts ss ON sa.sql_script_id = ss.id
            JOIN email_configs ec ON sa.email_config_id = ec.id
            WHERE sa.id = ?
        ''', (alert_id,))
        
        alert = cursor.fetchone()
        
        conn.close()
        
        if not alert:
            return jsonify({'error': '找不到指定的SQL预警'}), 404
        
        # 转换为字典并添加兼容字段
        alert_dict = dict(alert)
        alert_dict['emails'] = alert_dict['recipients']
        alert_dict['enabled'] = alert_dict['is_enabled']
        alert_dict['script_name'] = alert_dict['sql_script_name']
        alert_dict['script_id'] = alert_dict['sql_script_id']
        alert_dict['condition'] = alert_dict['condition_type']  # 添加condition字段映射
        # 添加threshold字段，默认为1
        alert_dict['threshold'] = alert_dict.get('threshold', 1)
        # 添加cron_expression和next_check字段
        alert_dict['cron_expression'] = alert_dict.get('cron_expression')
        alert_dict['next_check'] = alert_dict.get('next_check')
        
        return jsonify(alert_dict)
    except Exception as e:
        logger.error(f"获取SQL预警失败: {e}")
        return jsonify({'error': '获取SQL预警失败'}), 500

@app.route('/api/sql-alerts/<int:alert_id>', methods=['PUT'])
@login_required
def update_sql_alert(alert_id):
    """更新SQL查询预警"""
    logger.info(f"开始更新SQL预警 {alert_id}")
    try:
        data = request.get_json()
        logger.info(f"请求数据: {data}")
        
        # 支持前端发送的字段名
        name = data.get('name')
        sql_script_id = data.get('sql_script_id') or data.get('script_id')
        email_config_id = data.get('email_config_id')
        recipients = data.get('recipients') or data.get('emails')
        condition_type = data.get('condition_type') or data.get('condition')
        threshold = data.get('threshold')
        is_enabled = data.get('is_enabled') or data.get('enabled')
        cron_expression = data.get('cron_expression')
        
        logger.info(f"解析后的字段 - name: {name}, is_enabled: {is_enabled}, sql_script_id: {sql_script_id}")
        
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        
        # 检查sql_alerts表是否有cron_expression和next_check字段，如果没有则添加
        cursor.execute("PRAGMA table_info(sql_alerts)")
        sql_alert_columns = [column[1] for column in cursor.fetchall()]
        
        if 'cron_expression' not in sql_alert_columns:
            logger.info("Adding cron_expression column to sql_alerts table")
            cursor.execute("ALTER TABLE sql_alerts ADD COLUMN cron_expression TEXT")
            logger.info("Added cron_expression column to sql_alerts table")
        
        if 'next_check' not in sql_alert_columns:
            logger.info("Adding next_check column to sql_alerts table")
            cursor.execute("ALTER TABLE sql_alerts ADD COLUMN next_check TIMESTAMP")
            logger.info("Added next_check column to sql_alerts table")
        
        # 构建更新语句
        update_fields = []
        update_values = []
        
        if name is not None:
            update_fields.append('name = ?')
            update_values.append(name)
        
        if sql_script_id is not None:
            update_fields.append('sql_script_id = ?')
            update_values.append(sql_script_id)
        
        if email_config_id is not None:
            update_fields.append('email_config_id = ?')
            update_values.append(email_config_id)
        
        if recipients is not None:
            update_fields.append('recipients = ?')
            update_values.append(recipients)
        
        if condition_type is not None:
            update_fields.append('condition_type = ?')
            update_values.append(condition_type)
        
        if threshold is not None:
            # 检查是否有threshold字段，如果没有则添加
            cursor.execute("PRAGMA table_info(sql_alerts)")
            columns = [column[1] for column in cursor.fetchall()]
            if 'threshold' not in columns:
                cursor.execute("ALTER TABLE sql_alerts ADD COLUMN threshold INTEGER DEFAULT 1")
            
            update_fields.append('threshold = ?')
            update_values.append(threshold)
        
        if is_enabled is not None:
            # 检查当前预警状态
            cursor.execute('SELECT is_enabled FROM sql_alerts WHERE id = ?', (alert_id,))
            current_status = cursor.fetchone()
            logger.info(f"SQL预警 {alert_id} 当前状态: {current_status[0] if current_status else None}, 请求状态: {is_enabled}")
            
            # 只有当状态不同时才更新
            if current_status and current_status[0] != is_enabled:
                update_fields.append('is_enabled = ?')
                update_values.append(is_enabled)
                logger.info(f"SQL预警 {alert_id} 状态将从 {current_status[0]} 更新为 {is_enabled}")
            else:
                logger.info(f"SQL预警 {alert_id} 状态未改变，不需要更新")
        
        # 处理cron表达式和下次检查时间
        if cron_expression is not None:
            update_fields.append('cron_expression = ?')
            update_values.append(cron_expression)
            
            # 如果提供了cron表达式，计算下次检查时间
            if cron_expression.strip():
                try:
                    from croniter import croniter
                    cron = croniter(cron_expression, datetime.now())
                    next_check = cron.get_next(datetime)
                    update_fields.append('next_check = ?')
                    update_values.append(next_check)
                    logger.info(f"SQL预警 {name or alert_id} 下次检查时间: {next_check}")
                except Exception as e:
                    logger.error(f"Invalid cron expression for SQL alert {name or alert_id}: {cron_expression}, error: {e}")
                    return jsonify({'error': f'无效的cron表达式: {cron_expression}'}), 400
            else:
                # 如果cron表达式为空，清空下次检查时间
                update_fields.append('next_check = ?')
                update_values.append(None)
        
        if not update_fields:
            conn.close()
            # 如果没有字段需要更新，返回成功响应
            return jsonify({'success': True, 'message': '状态未改变'})
        
        update_fields.append('updated_at = CURRENT_TIMESTAMP')
        update_values.append(alert_id)
        
        cursor.execute(f'''
            UPDATE sql_alerts
            SET {', '.join(update_fields)}
            WHERE id = ?
        ''', tuple(update_values))
        
        conn.commit()
        
        # 获取更新后的预警信息
        cursor.execute('SELECT * FROM sql_alerts WHERE id = ?', (alert_id,))
        alert = cursor.fetchone()
        conn.close()
        
        if not alert:
            return jsonify({'error': 'SQL预警不存在'}), 404
        
        return jsonify({
            'success': True,
            'message': 'SQL预警更新成功',
            'alert': {
                'id': alert[0],
                'name': alert[1],
                'sql_script_id': alert[2],
                'email_config_id': alert[3],
                'recipients': alert[4],
                'condition_type': alert[5],
                'threshold': alert[10],
                'is_enabled': alert[6],
                'last_check': alert[7],
                'created_at': alert[8],
                'updated_at': alert[9],
                'cron_expression': alert[11] if len(alert) > 11 else None,
                'next_check': alert[12] if len(alert) > 12 else None
            }
        })
    except Exception as e:
        logger.error(f"更新SQL预警失败: {str(e)}")
        return jsonify({'error': f'更新SQL预警失败: {str(e)}'}), 500

@app.route('/api/sql-alerts/<int:alert_id>', methods=['DELETE'])
@login_required
def delete_sql_alert(alert_id):
    """删除SQL查询预警"""
    try:
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        
        cursor.execute('DELETE FROM sql_alerts WHERE id = ?', (alert_id,))
        
        if cursor.rowcount == 0:
            conn.close()
            return jsonify({'error': '预警不存在'}), 404
        
        conn.commit()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': 'SQL预警删除成功'
        })
    except Exception as e:
        logger.error(f"删除SQL预警失败: {str(e)}")
        return jsonify({'error': '删除SQL预警失败'}), 500

@app.route('/api/sql-alerts/<int:alert_id>/check', methods=['POST'])
@login_required
def check_sql_alert(alert_id):
    """手动检查SQL预警"""
    result = _check_sql_alert_internal(alert_id)
    
    # 将内部结果转换为Flask响应
    if isinstance(result, dict) and 'error' in result:
        status_code = 400
        if '不存在' in result['error']:
            status_code = 404
        elif '执行SQL查询失败' in result['error'] or '检查SQL预警失败' in result['error']:
            status_code = 500
        return jsonify(result), status_code
    else:
        return jsonify(result)


def _check_sql_alert_internal(alert_id):
    """内部SQL预警检查函数，不依赖Flask请求上下文"""
    # 创建日志收集器
    log_collector = LogCollector()
    
    try:
        # 开始收集日志
        log_collector.start_collecting()
        logger.info(f"开始检查SQL预警，ID: {alert_id}")
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        
        # 获取预警配置
        cursor.execute('''
            SELECT sa.id, sa.name, sa.sql_script_id, sa.email_config_id, sa.recipients, 
                   sa.condition_type, sa.is_enabled, sa.last_check, sa.created_at, 
                   sa.updated_at, sa.threshold, sa.cron_expression, sa.next_check,
                   ss.sql_content, ss.db_config_id,
                   ec.mail_server, ec.mail_port, ec.mail_use_tls, 
                   ec.mail_username, ec.mail_password, ec.mail_default_sender
            FROM sql_alerts sa
            JOIN sql_scripts ss ON sa.sql_script_id = ss.id
            JOIN email_configs ec ON sa.email_config_id = ec.id
            WHERE sa.id = ?
        ''', (alert_id,))
        
        alert_config = cursor.fetchone()
        if not alert_config:
            conn.close()
            logger.error(f"SQL预警不存在，ID: {alert_id}")
            return {'error': '预警不存在'}
        
        # 使用正确的索引位置获取字段值
        alert_id = alert_config[0]
        name = alert_config[1]
        sql_script_id = alert_config[2]
        email_config_id = alert_config[3]
        recipients = alert_config[4]
        condition_type = alert_config[5]
        is_enabled = alert_config[6]
        last_check = alert_config[7]
        created_at = alert_config[8]
        updated_at = alert_config[9]
        threshold = alert_config[10]
        cron_expression = alert_config[11]
        next_check = alert_config[12]
        # sql_scripts表的字段从索引13开始
        script_content = alert_config[13]
        db_config_id = alert_config[14]  # 修正：使用正确的索引位置
        # email_configs表的字段从索引15开始
        mail_server = alert_config[15]
        mail_port = alert_config[16]
        mail_use_tls = alert_config[17]
        mail_username = alert_config[18]
        mail_password = alert_config[19]
        mail_default_sender = alert_config[20]
        
        logger.info(f"SQL预警配置: 名称={name}, 条件类型={condition_type}, 启用状态={is_enabled}")
        
        if not is_enabled:
            conn.close()
            logger.error(f"SQL预警已禁用，ID: {alert_id}")
            return {'error': '预警已禁用'}
        
        # 获取数据库连接配置 (使用SQL脚本关联的配置)
        # 检查db_config_id是否为空
        if not db_config_id:
            conn.close()
            logger.error(f'SQL脚本没有关联的数据库配置 (SQL脚本ID: {sql_script_id})')
            return {'error': f'SQL脚本没有关联的数据库配置 (SQL脚本ID: {sql_script_id})'}
            
        cursor.execute('SELECT * FROM db_configs WHERE id = ?', (db_config_id,))
        db_config = cursor.fetchone()
        
        if not db_config:
            conn.close()
            logger.error(f'SQL脚本关联的数据库配置不存在 (ID: {db_config_id})')
            return {'error': f'SQL脚本关联的数据库配置不存在 (ID: {db_config_id})'}
        
        # 使用索引而不是解包，因为db_configs表可能有额外的字段
        db_id = db_config[0]
        db_name = db_config[1]
        db_type = db_config[2]
        host = db_config[3]
        port = db_config[4]
        username = db_config[5]
        password = db_config[6]
        created_at = db_config[7]
        is_default = db_config[8] if len(db_config) > 8 else None
        
        logger.info(f"数据库配置: 类型={db_type}, 主机={host}, 端口={port}, 用户名={username}")
        
        # 连接数据库并执行SQL
        try:
            logger.info(f"尝试连接数据库...")
            # 从SQL脚本中提取数据库名（如果存在）
            database_name = None
            table_name = None
            if db_type.lower() == 'mysql':
                # 简单的正则表达式来提取数据库名和表名
                import re
                # 查找 "from database.table" 或 "join database.table" 模式
                matches = re.findall(r'\b(?:from|join)\s+([a-zA-Z_][a-zA-Z0-9_]*)\.([a-zA-Z_][a-zA-Z0-9_]*)', script_content, re.IGNORECASE)
                if matches:
                    database_name = matches[0][0]  # 数据库名
                    table_name = matches[0][1]     # 表名
                    logger.info(f"从SQL中提取到数据库名: {database_name}, 表名: {table_name}")
                
                import pymysql
                
                # 首先尝试使用提取的数据库名连接
                try:
                    logger.info(f"尝试连接到MySQL数据库: {host}:{port}, 数据库: {database_name}")
                    db_conn = pymysql.connect(
                        host=host,
                        port=int(port),
                        user=username,
                        password=password,
                        database=database_name,
                        charset='utf8mb4'
                    )
                    logger.info("MySQL连接成功")
                except Exception as first_error:
                    logger.error(f"使用数据库名 {database_name} 连接失败: {str(first_error)}")
                    logger.error(f"连接失败详细信息: 主机={host}, 端口={port}, 用户名={username}, 数据库={database_name}")
                    logger.error(f"完整错误信息: {repr(first_error)}")
                    # 如果连接失败，尝试在所有可用数据库中查找表
                    if table_name:
                        logger.info(f"尝试在所有数据库中查找表 {table_name}")
                        try:
                            temp_conn = pymysql.connect(
                                host=host,
                                port=int(port),
                                user=username,
                                password=password,
                                charset='utf8mb4'
                            )
                            logger.info(f"成功连接到MySQL服务器 {host}:{port}，开始查找数据库")
                        except Exception as conn_error:
                            logger.error(f"无法连接到MySQL服务器 {host}:{port}: {str(conn_error)}")
                            logger.error(f"连接错误详细信息: {repr(conn_error)}")
                            raise Exception(f"无法连接到MySQL服务器 {host}:{port}: {str(conn_error)}")
                            
                        temp_cursor = temp_conn.cursor()
                        
                        # 获取所有数据库列表
                        try:
                            temp_cursor.execute("SHOW DATABASES")
                            all_databases = [db[0] for db in temp_cursor.fetchall()]
                            logger.info(f"服务器上共有 {len(all_databases)} 个数据库")
                        except Exception as db_list_error:
                            logger.error(f"获取数据库列表失败: {str(db_list_error)}")
                            temp_conn.close()
                            raise Exception(f"获取数据库列表失败: {str(db_list_error)}")
                        
                        # 排除系统数据库
                        user_databases = [db for db in all_databases if not db.startswith('information_schema') and not db.startswith('mysql') and not db.startswith('__')]
                        logger.info(f"找到 {len(user_databases)} 个用户数据库: {user_databases}")
                        
                        # 在每个用户数据库中查找表
                        found_database = None
                        for db in user_databases:
                            try:
                                temp_cursor.execute(f"SHOW TABLES FROM `{db}` LIKE '{table_name}'")
                                if temp_cursor.fetchone():
                                    found_database = db
                                    logger.info(f"在数据库 {db} 中找到表 {table_name}")
                                    break
                            except Exception as table_check_error:
                                logger.warning(f"检查数据库 {db} 中的表时出错: {str(table_check_error)}")
                                continue
                        
                        temp_conn.close()
                        
                        if found_database:
                            # 使用找到的数据库连接
                            logger.info(f"使用找到的数据库 {found_database} 连接")
                            try:
                                db_conn = pymysql.connect(
                                    host=host,
                                    port=int(port),
                                    user=username,
                                    password=password,
                                    database=found_database,
                                    charset='utf8mb4'
                                )
                                logger.info(f"成功连接到数据库 {found_database}")
                            except Exception as final_conn_error:
                                logger.error(f"连接到数据库 {found_database} 失败: {str(final_conn_error)}")
                                logger.error(f"最终连接错误详细信息: {repr(final_conn_error)}")
                                raise Exception(f"连接到数据库 {found_database} 失败: {str(final_conn_error)}")
                                
                            # 修改SQL脚本中的数据库名
                            script_content = script_content.replace(f"{database_name}.{table_name}", f"{found_database}.{table_name}")
                            logger.info(f"已修改SQL中的数据库名从 {database_name} 到 {found_database}")
                        else:
                            logger.error(f"在所有数据库中都找不到表 {table_name}")
                            logger.error(f"已检查的数据库: {user_databases}")
                            raise Exception(f"在所有数据库中都找不到表 {table_name}")
                    else:
                        logger.error(f"无法从SQL中提取表名，无法进行数据库查找")
                        raise first_error
            else:  # SQLite
                logger.info("连接到SQLite数据库")
                db_conn = sqlite3.connect('scheduler.db')
            
            logger.info(f"执行SQL查询...")
            logger.info(f"SQL内容: {script_content[:200]}{'...' if len(script_content) > 200 else ''}")
            cursor_db = db_conn.cursor()
            try:
                cursor_db.execute(script_content)
                results = cursor_db.fetchall()
                
                # 获取列名
                column_names = None
                if cursor_db.description:
                    column_names = [desc[0] for desc in cursor_db.description]
                    logger.info(f"SQL列名: {column_names}")
                
                logger.info(f"SQL执行成功，返回 {len(results)} 条结果")
                if results and len(results) > 0:
                    logger.info(f"第一行结果示例: {results[0][:5]}{'...' if len(results[0]) > 5 else ''}")
            except Exception as sql_error:
                logger.error(f"SQL执行失败: {str(sql_error)}")
                logger.error(f"SQL错误详细信息: {repr(sql_error)}")
                logger.error(f"执行的SQL: {script_content}")
                logger.error(f"数据库连接信息: {host}:{port}/{database_name if database_name else 'N/A'}")
                raise sql_error
            finally:
                try:
                    cursor_db.close()
                except:
                    pass
            
            db_conn.close()
            logger.info(f"SQL执行成功，返回 {len(results)} 条结果")
            
            # 检查条件
            should_alert = False
            if condition_type == 'not_empty':
                should_alert = len(results) > 0
            elif condition_type == 'rows_gt':
                should_alert = len(results) > threshold
            elif condition_type == 'rows_lt':
                should_alert = len(results) < threshold
            elif condition_type == 'rows_eq':
                should_alert = len(results) == threshold
            elif condition_type == 'rows_neq':
                should_alert = len(results) != threshold
            
            logger.info(f"条件检查: 条件类型={condition_type}, 结果数量={len(results)}, 是否触发预警={should_alert}")
            
            if should_alert:
                # 发送预警邮件
                subject = f"SQL预警: {name}"
                body = f"SQL查询触发了预警通知。\n\n"
                body += f"预警名称: {name}\n"
                body += f"SQL脚本ID: {sql_script_id}\n"
                body += f"查询结果数: {len(results)} 条\n"
                body += f"检查时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n"
                
                # 创建Excel文件作为附件
                attachments = []
                excel_file = create_excel_file(results, name, column_names)
                if excel_file:
                    attachments.append(excel_file)
                    body += f"\n详细查询结果已导出为Excel文件，请查看附件。"
                    logger.info(f"Excel附件已添加: {excel_file}")
                else:
                    logger.warning("Excel文件创建失败，邮件将不包含附件")
                
                # 更新最后检查时间
                cursor.execute('UPDATE sql_alerts SET last_check = CURRENT_TIMESTAMP WHERE id = ?', (alert_id,))
                conn.commit()
                
                logger.info(f"发送预警邮件到: {recipients}")
                if attachments:
                    logger.info(f"邮件包含 {len(attachments)} 个附件")
                else:
                    logger.warning("邮件不包含任何附件")
                
                # 添加邮件配置日志记录（不包含密码）
                logger.info(f"邮件配置信息:")
                logger.info(f"  - 邮件服务器: {mail_server}")
                logger.info(f"  - 邮件端口: {mail_port}")
                logger.info(f"  - 使用TLS: {mail_use_tls}")
                logger.info(f"  - 用户名: {mail_username}")
                logger.info(f"  - 默认发件人: {mail_default_sender}")
                logger.info(f"  - 收件人: {recipients}")
                    
                # 发送邮件
                email_sent = True
                try:
                    send_alert_email(
                        mail_server, mail_port, mail_use_tls, mail_username, mail_password, mail_default_sender,
                        recipients, subject, body, attachments
                    )
                    logger.info("预警邮件发送成功")
                except Exception as email_error:
                    email_sent = False
                    logger.error(f"预警邮件发送失败: {str(email_error)}")
                    # 不抛出异常，继续执行其他逻辑
                
                # 清理临时文件
                if excel_file and os.path.exists(excel_file):
                    try:
                        os.unlink(excel_file)
                    except:
                        pass
                
                conn.close()
                logger.info("SQL预警处理完成")
                
                # 记录日志
                if email_sent:
                    log_status = 'triggered'
                    log_message = f"SQL预警触发，查询结果数: {len(results)} 条，邮件发送成功"
                else:
                    log_status = 'triggered_email_failed'
                    log_message = f"SQL预警触发，查询结果数: {len(results)} 条，邮件发送失败"
                
                # 创建详细日志信息
                execution_logs = log_collector.get_collected_logs()
                details = f"""SQL预警执行详细信息:
预警名称: {name}
SQL脚本ID: {sql_script_id}
数据库配置ID: {db_config_id}
数据库类型: {db_type}
主机: {host}
端口: {port}
条件类型: {condition_type}
阈值: {threshold}
查询结果数: {len(results)}
检查时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
邮件发送状态: {'成功' if email_sent else '失败'}
收件人: {recipients}

SQL脚本:
{script_content}

查询结果:
"""
                # 添加查询结果到详细信息中（限制前10条以避免日志过长）
                if results:
                    details += f"列名: {', '.join(column_names) if column_names else '未知'}\n"
                    details += f"总行数: {len(results)}\n"
                    details += "前10行数据:\n"
                    for i, row in enumerate(results[:10]):
                        details += f"  行{i+1}: {row}\n"
                    if len(results) > 10:
                        details += f"... 还有 {len(results)-10} 行数据\n"
                else:
                    details += "查询结果为空\n"
                
                # 添加执行过程日志
                details += f"\n执行过程日志:\n{execution_logs}"
                
                log_sql_alert_execution(alert_id, name, log_status, log_message, details)
                
                return {
                    'success': True,
                    'message': f'SQL预警触发，已发送邮件通知',
                    'result_count': len(results),
                    'triggered': True,
                    'condition_text': get_condition_text(condition_type, threshold),
                    'email_sent': email_sent
                }
            else:
                # 更新最后检查时间
                cursor.execute('UPDATE sql_alerts SET last_check = CURRENT_TIMESTAMP WHERE id = ?', (alert_id,))
                conn.commit()
                conn.close()
                logger.info("SQL检查完成，未触发预警")
                
                # 记录日志
                log_message = f"SQL检查完成，查询结果数: {len(results)} 条，未触发预警条件"
                
                # 创建详细日志信息
                execution_logs = log_collector.get_collected_logs()
                details = f"""SQL预警执行详细信息:
预警名称: {name}
SQL脚本ID: {sql_script_id}
数据库配置ID: {db_config_id}
数据库类型: {db_type}
主机: {host}
端口: {port}
条件类型: {condition_type}
阈值: {threshold}
查询结果数: {len(results)}
检查时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
触发状态: 未触发预警条件

SQL脚本:
{script_content}

查询结果:
"""
                # 添加查询结果到详细信息中（限制前10条以避免日志过长）
                if results:
                    details += f"列名: {', '.join(column_names) if column_names else '未知'}\n"
                    details += f"总行数: {len(results)}\n"
                    details += "前10行数据:\n"
                    for i, row in enumerate(results[:10]):
                        details += f"  行{i+1}: {row}\n"
                    if len(results) > 10:
                        details += f"... 还有 {len(results)-10} 行数据\n"
                else:
                    details += "查询结果为空\n"
                
                # 添加执行过程日志
                details += f"\n执行过程日志:\n{execution_logs}"
                
                log_sql_alert_execution(alert_id, name, 'no_trigger', log_message, details)
                
                return {
                    'success': True,
                    'message': 'SQL检查完成，未触发预警',
                    'result_count': len(results),
                    'triggered': False,
                    'condition_text': get_condition_text(condition_type, threshold)
                }
                
        except Exception as e:
            conn.close()
            logger.error(f"执行SQL查询失败: {str(e)}")
            logger.error(f"SQL预警详细信息:")
            logger.error(f"  - 预警ID: {alert_id}")
            logger.error(f"  - 预警名称: {name}")
            logger.error(f"  - SQL脚本ID: {sql_script_id}")
            logger.error(f"  - 数据库配置ID: {db_config_id}")
            logger.error(f"  - 数据库类型: {db_type}")
            logger.error(f"  - 主机: {host}")
            logger.error(f"  - 端口: {port}")
            logger.error(f"  - 用户名: {username}")
            logger.error(f"  - 条件类型: {condition_type}")
            logger.error(f"  - 阈值: {threshold}")
            import traceback
            logger.error(f"SQL查询错误堆栈: {traceback.format_exc()}")
            
            # 记录日志
            log_message = f"执行SQL查询失败: {str(e)}"
            
            # 创建详细日志信息
            execution_logs = log_collector.get_collected_logs()
            details = f"""SQL预警执行详细信息:
预警名称: {name}
SQL脚本ID: {sql_script_id}
数据库配置ID: {db_config_id}
数据库类型: {db_type}
主机: {host}
端口: {port}
条件类型: {condition_type}
阈值: {threshold}
检查时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
执行状态: 失败

SQL脚本:
{script_content}

错误信息:
{str(e)}

错误堆栈:
{traceback.format_exc()}

执行过程日志:
{execution_logs}
"""
            
            log_sql_alert_execution(alert_id, name, 'failed', log_message, details)
            
            return {'error': f'执行SQL查询失败: {str(e)}'}
            
    except Exception as e:
        logger.error(f"检查SQL预警失败: {str(e)}")
        logger.error(f"请求参数: alert_id={alert_id}")
        import traceback
        logger.error(f"检查SQL预警错误堆栈: {traceback.format_exc()}")
        
        # 记录日志
        log_message = f"检查SQL预警失败: {str(e)}"
        
        # 创建详细日志信息
        details = f"""SQL预警执行详细信息:
预警ID: {alert_id}
检查时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
执行状态: 失败

错误信息:
{str(e)}

错误堆栈:
{traceback.format_exc()}
"""
        
        log_sql_alert_execution(alert_id, "未知预警", 'failed', log_message, details)
        
        return {'error': f'检查SQL预警失败: {str(e)}'}

def create_excel_file(results, alert_name, column_names=None):
    """将查询结果导出为Excel文件"""
    if pd is None or openpyxl is None:
        logger.error("pandas或openpyxl未安装，无法生成Excel文件")
        return None
        
    try:
        # 获取列名（假设results是元组列表）
        if not results:
            logger.warning("查询结果为空，不创建Excel文件")
            return None
            
        # 创建DataFrame，使用提供的列名作为表头
        if column_names:
            df = pd.DataFrame(results, columns=column_names)
        else:
            df = pd.DataFrame(results)
        
        # 创建临时文件，使用预警名称作为文件名
        # 清理文件名中的非法字符
        import re
        from datetime import datetime, date
        safe_name = re.sub(r'[\\/*?:"<>|]', "", alert_name)  # 移除Windows文件名中的非法字符
        safe_name = safe_name.replace(" ", "_")  # 替换空格为下划线
        
        # 创建临时目录和文件
        temp_dir = tempfile.gettempdir()
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        file_name = f"{safe_name}_{timestamp}.xlsx"
        file_path = os.path.join(temp_dir, file_name)
        
        logger.info(f"开始创建Excel文件: {file_path}")
        logger.info(f"数据行数: {len(df)}, 数据列数: {len(df.columns)}")
        
        # 使用ExcelWriter进行更精细的控制
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='查询结果')
            
            # 获取工作表
            worksheet = writer.sheets['查询结果']
            
            # 遍历所有列，检测可能的日期列
            for col_num, column in enumerate(df.columns):
                # 检查列名是否包含日期相关关键词
                if any(keyword in column.lower() for keyword in ['date', 'time', '日期', '时间', '创建', '更新', '开始', '结束']):
                    # 设置列宽
                    try:
                        if get_column_letter:
                            worksheet.column_dimensions[get_column_letter(col_num + 1)].width = 20
                    except Exception as col_error:
                        logger.warning(f"设置列宽失败: {str(col_error)}")
                    
                    # 遍历该列的所有单元格
                    for row_num in range(2, len(df) + 2):  # 从第2行开始（第1行是表头）
                        cell = worksheet.cell(row=row_num, column=col_num + 1)
                        
                        # 尝试解析日期
                        if cell.value and isinstance(cell.value, str):
                            # 尝试匹配常见日期格式
                            date_patterns = [
                                r'\d{4}-\d{2}-\d{2}',  # YYYY-MM-DD
                                r'\d{4}/\d{2}/\d{2}',  # YYYY/MM/DD
                                r'\d{2}-\d{2}-\d{4}',  # MM-DD-YYYY
                                r'\d{2}/\d{2}/\d{4}',  # MM/DD/YYYY
                                r'\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}',  # YYYY-MM-DD HH:MM:SS
                                r'\d{4}/\d{2}/\d{2} \d{2}:\d{2}:\d{2}',  # YYYY/MM/DD HH:MM:SS
                            ]
                            
                            for pattern in date_patterns:
                                if re.match(pattern, cell.value):
                                    try:
                                        # 尝试解析为日期
                                        if ' ' in cell.value:  # 包含时间
                                            parsed_date = datetime.strptime(cell.value, '%Y-%m-%d %H:%M:%S') if '-' in cell.value else datetime.strptime(cell.value, '%Y/%m/%d %H:%M:%S')
                                            cell.number_format = 'YYYY-MM-DD HH:MM:SS'
                                        else:  # 只有日期
                                            parsed_date = datetime.strptime(cell.value, '%Y-%m-%d') if '-' in cell.value else datetime.strptime(cell.value, '%Y/%m/%d')
                                            cell.number_format = 'YYYY-MM-DD'
                                        
                                        cell.value = parsed_date
                                        break
                                    except:
                                        # 如果解析失败，保持原值
                                        pass
                        elif cell.value and isinstance(cell.value, (datetime, date)):
                            # 如果已经是日期类型，设置格式
                            if isinstance(cell.value, datetime):
                                cell.number_format = 'YYYY-MM-DD HH:MM:SS'
                            else:
                                cell.number_format = 'YYYY-MM-DD'
        
        # 验证文件是否创建成功
        if os.path.exists(file_path) and os.path.getsize(file_path) > 0:
            logger.info(f"Excel文件已成功创建: {file_path}, 文件大小: {os.path.getsize(file_path)} 字节")
            return file_path
        else:
            logger.error(f"Excel文件创建失败或文件为空: {file_path}")
            return None
            
    except Exception as e:
        logger.error(f"创建Excel文件失败: {str(e)}")
        import traceback
        logger.error(f"详细错误信息: {traceback.format_exc()}")
        return None

def send_alert_email(mail_server, mail_port, mail_use_tls, mail_username, mail_password, mail_default_sender, recipients, subject, body, attachments=None):
    """发送预警邮件"""
    try:
        # 添加邮件配置日志记录（不包含密码）
        logger.info(f"邮件发送函数 - 邮件配置信息:")
        logger.info(f"  - 邮件服务器: {mail_server}")
        logger.info(f"  - 邮件端口: {mail_port}")
        logger.info(f"  - 使用TLS: {mail_use_tls}")
        logger.info(f"  - 用户名: {mail_username}")
        logger.info(f"  - 默认发件人: {mail_default_sender}")
        logger.info(f"  - 收件人: {recipients}")
        
        # 确保在应用上下文中执行
        with app.app_context():
            # 配置邮件
            app.config['MAIL_SERVER'] = mail_server
            app.config['MAIL_PORT'] = mail_port
            app.config['MAIL_USERNAME'] = mail_username
            app.config['MAIL_PASSWORD'] = mail_password
            app.config['MAIL_USE_TLS'] = mail_use_tls
            app.config['MAIL_DEFAULT_SENDER'] = mail_default_sender or mail_username
            
            logger.info("邮件配置已设置到Flask应用")
            
            # 重新初始化邮件对象以应用新配置
            global mail
            mail = Mail(app)
            logger.info("邮件对象已重新初始化")
            
            # 解析收件人列表
            recipient_list = [r.strip() for r in recipients.split(',')]
            logger.info(f"准备发送邮件给 {len(recipient_list)} 个收件人: {recipient_list}")
            
            # 发送邮件给每个收件人
            for recipient in recipient_list:
                logger.info(f"创建邮件对象，收件人: {recipient}")
                msg = Message(
                    subject=subject,
                    recipients=[recipient],
                    body=body,
                    sender=mail_default_sender or mail_username
                )
                
                # 添加附件
                if attachments:
                    logger.info(f"处理 {len(attachments)} 个附件")
                    for i, attachment_path in enumerate(attachments):
                        logger.info(f"处理附件 {i+1}/{len(attachments)}: {attachment_path}")
                        if os.path.exists(attachment_path):
                            file_size = os.path.getsize(attachment_path)
                            logger.info(f"附件文件存在，大小: {file_size} 字节")
                            with app.open_resource(attachment_path) as fp:
                                msg.attach(
                                    filename=os.path.basename(attachment_path),
                                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                                    data=fp.read()
                                )
                            logger.info(f"附件 {os.path.basename(attachment_path)} 已添加到邮件")
                        else:
                            logger.error(f"附件文件不存在: {attachment_path}")
                else:
                    logger.warning("没有附件需要添加到邮件")
                
                logger.info(f"发送邮件给 {recipient}")
                try:
                    mail.send(msg)
                    logger.info(f"邮件已成功发送给 {recipient}")
                except Exception as send_error:
                    logger.error(f"发送邮件给 {recipient} 失败: {str(send_error)}")
                    raise send_error
            
            logger.info(f"预警邮件已发送至: {', '.join(recipient_list)}")
    except Exception as e:
        logger.error(f"发送预警邮件失败: {str(e)}")
        import traceback
        logger.error(f"详细错误信息: {traceback.format_exc()}")
        raise


@app.route('/api/db-configs/<int:config_id>', methods=['GET'])
@login_required
def get_db_config(config_id):
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM db_configs WHERE id = ?', (config_id,))
    config = cursor.fetchone()
    conn.close()

    if not config:
        return jsonify({'error': '数据库配置不存在'}), 404

    config_dict = {
        'id': config[0],
        'name': config[1],
        'db_type': config[2],
        'host': config[3],
        'port': config[4],
        'username': config[5],
        'password': config[6],
        'created_at': config[7]
    }

    return jsonify(config_dict)


# 数据库配置API
@app.route('/api/db-configs', methods=['GET'])
@login_required
def get_db_configs():
    logger.info("API call: get_db_configs - Starting function")
    logger.info(f"Request method: {request.method}")
    logger.info(f"Request path: {request.path}")
    logger.info(f"Session data: {dict(session)}")
    
    try:
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        cursor.execute('SELECT * FROM db_configs ORDER BY id DESC')
        configs = cursor.fetchall()
        conn.close()
        
        logger.info(f"Retrieved {len(configs)} db configs from database")
        
        configs_list = []
        for config in configs:
            configs_list.append({
                'id': config[0],
                'name': config[1],
                'db_type': config[2],
                'host': config[3],
                'port': config[4],
                'username': config[5],
                'password': config[6],
                'created_at': config[7]
            })

        logger.info(f"Processed {len(configs_list)} db configs, returning JSON response")
        return jsonify(configs_list)
        
    except Exception as e:
        logger.error(f"Error in get_db_configs: {str(e)}")
        return jsonify({'error': 'Internal server error'}), 500

@app.route('/api/db-configs', methods=['POST'])
@login_required
def create_db_config():
    data = request.json
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()

    # 使用北京时间记录创建时间
    created_at = get_beijing_time().isoformat()

    cursor.execute('''
        INSERT INTO db_configs (name, db_type, host, port, username, password, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?)
    ''', (
        data['name'],
        data['db_type'],
        data.get('host'),
        data.get('port'),
        data.get('username'),
        data.get('password'),
        created_at
    ))

    config_id = cursor.lastrowid
    conn.commit()
    conn.close()

    return jsonify({'message': '数据库配置创建成功', 'id': config_id})


@app.route('/api/db-configs/<int:config_id>', methods=['PUT'])
@login_required
def update_db_config(config_id):
    data = request.json
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()

    cursor.execute('''
        UPDATE db_configs 
        SET name=?, db_type=?, host=?, port=?, username=?, password=?
        WHERE id=?
    ''', (
        data['name'],
        data['db_type'],
        data.get('host'),
        data.get('port'),
        data.get('username'),
        data.get('password'),
        config_id
    ))

    conn.commit()
    conn.close()

    return jsonify({'message': '数据库配置更新成功'})


@app.route('/api/db-configs/<int:config_id>', methods=['DELETE'])
@login_required
def delete_db_config(config_id):
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()

    # 检查是否有SQL脚本正在使用这个配置
    cursor.execute('SELECT COUNT(*) FROM sql_scripts WHERE db_config_id = ?', (config_id,))
    count = cursor.fetchone()[0]

    if count > 0:
        conn.close()
        return jsonify({'error': f'有{count}个SQL脚本正在使用此数据库配置，无法删除'}), 400

    cursor.execute('DELETE FROM db_configs WHERE id = ?', (config_id,))
    conn.commit()
    conn.close()

    return jsonify({'message': '数据库配置删除成功'})


# SQL脚本API
@app.route('/api/sql-scripts', methods=['GET'])
@login_required
def get_sql_scripts():
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute('''
        SELECT s.*, c.name as db_config_name
        FROM sql_scripts s
        LEFT JOIN db_configs c ON s.db_config_id = c.id
        ORDER BY s.id DESC
    ''')
    scripts = cursor.fetchall()
    conn.close()

    scripts_list = []
    for script in scripts:
        scripts_list.append({
            'id': script[0],
            'name': script[1],
            'db_config_id': script[2],
            'sql_content': script[3],
            'created_at': script[4],
            'db_config_name': script[5]
        })

    return jsonify({'sql_scripts': scripts_list})


@app.route('/api/sql-scripts/<int:script_id>', methods=['GET'])
@login_required
def get_sql_script(script_id):
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute('''
        SELECT s.*, c.name as db_config_name
        FROM sql_scripts s
        LEFT JOIN db_configs c ON s.db_config_id = c.id
        WHERE s.id = ?
    ''', (script_id,))
    script = cursor.fetchone()
    conn.close()

    if not script:
        return jsonify({'error': 'SQL脚本不存在'}), 404

    script_dict = {
        'id': script[0],
        'name': script[1],
        'db_config_id': script[2],
        'sql_content': script[3],
        'created_at': script[4],
        'db_config_name': script[5]
    }

    return jsonify(script_dict)


@app.route('/api/sql-scripts', methods=['POST'])
@login_required
def create_sql_script():
    data = request.json
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()

    # 使用北京时间记录创建时间
    created_at = get_beijing_time().isoformat()

    cursor.execute('''
        INSERT INTO sql_scripts (name, db_config_id, sql_content, created_at)
        VALUES (?, ?, ?, ?)
    ''', (
        data['name'],
        data['db_config_id'],
        data['sql_content'],
        created_at
    ))

    script_id = cursor.lastrowid
    conn.commit()
    conn.close()

    return jsonify({'message': 'SQL脚本创建成功', 'id': script_id})


@app.route('/api/sql-scripts/<int:script_id>', methods=['PUT'])
@login_required
def update_sql_script(script_id):
    data = request.json
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()

    cursor.execute('''
        UPDATE sql_scripts 
        SET name=?, db_config_id=?, sql_content=?
        WHERE id=?
    ''', (
        data['name'],
        data['db_config_id'],
        data['sql_content'],
        script_id
    ))

    conn.commit()
    conn.close()

    return jsonify({'message': 'SQL脚本更新成功'})


@app.route('/api/sql-scripts/<int:script_id>', methods=['DELETE'])
@login_required
def delete_sql_script(script_id):
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute('DELETE FROM sql_scripts WHERE id = ?', (script_id,))
    conn.commit()
    conn.close()

    return jsonify({'message': 'SQL脚本删除成功'})


# 用户管理API
@app.route('/api/users', methods=['GET'])
@login_required
def get_users():
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute('SELECT id, username, created_at FROM users ORDER BY id')
    users = cursor.fetchall()
    conn.close()

    users_list = []
    for user in users:
        users_list.append({
            'id': user[0],
            'username': user[1],
            'created_at': user[2]
        })

    return jsonify(users_list)


@app.route('/api/users', methods=['POST'])
@login_required
def create_user():
    data = request.json
    username = data.get('username')
    password = data.get('password')

    if not username or not password:
        return jsonify({'error': '用户名和密码不能为空'}), 400

    # 检查用户名是否已存在
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute('SELECT id FROM users WHERE username = ?', (username,))
    existing_user = cursor.fetchone()

    if existing_user:
        conn.close()
        return jsonify({'error': '用户名已存在'}), 400

    # 创建新用户
    password_hash = generate_password_hash(password)
    created_at = get_beijing_time().isoformat()

    cursor.execute('''
        INSERT INTO users (username, password_hash, created_at)
        VALUES (?, ?, ?)
    ''', (username, password_hash, created_at))

    user_id = cursor.lastrowid
    conn.commit()
    conn.close()

    return jsonify({'message': '用户创建成功', 'id': user_id})


@app.route('/api/users/change-password', methods=['POST'])
@login_required
def change_password():
    data = request.json
    current_password = data.get('current_password')
    new_password = data.get('new_password')
    user_id = session.get('user_id')

    if not current_password or not new_password:
        return jsonify({'error': '当前密码和新密码不能为空'}), 400

    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute('SELECT password_hash FROM users WHERE id = ?', (user_id,))
    user = cursor.fetchone()

    if not user:
        conn.close()
        return jsonify({'error': '用户不存在'}), 404

    current_password_hash = hashlib.sha256(current_password.encode()).hexdigest()
    if user[0] != current_password_hash:
        conn.close()
        return jsonify({'error': '当前密码错误'}), 400

    new_password_hash = hashlib.sha256(new_password.encode()).hexdigest()
    cursor.execute('UPDATE users SET password_hash = ? WHERE id = ?', (new_password_hash, user_id))

    conn.commit()
    conn.close()
    return jsonify({'message': '密码修改成功'})


@app.route('/api/users/<int:user_id>', methods=['DELETE'])
@login_required
def delete_user(user_id):
    # 获取当前登录用户ID
    current_user_id = session.get('user_id')
    
    # 不允许删除当前登录用户
    if user_id == current_user_id:
        return jsonify({'error': '不能删除当前登录用户'}), 400
    
    # 检查用户是否存在
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute('SELECT username FROM users WHERE id = ?', (user_id,))
    user = cursor.fetchone()
    
    if not user:
        conn.close()
        return jsonify({'error': '用户不存在'}), 404
    
    # 删除用户相关的Token
    cursor.execute('DELETE FROM user_tokens WHERE user_id = ?', (user_id,))
    
    # 删除用户
    cursor.execute('DELETE FROM users WHERE id = ?', (user_id,))
    
    conn.commit()
    conn.close()
    
    return jsonify({'message': f'用户 {user[0]} 删除成功'})


@app.route('/api/users/tokens', methods=['GET'])
@login_required
def get_user_tokens():
    user_id = session.get('user_id')
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute('''
        SELECT id, name, token, expires_at, created_at 
        FROM user_tokens 
        WHERE user_id = ? 
        ORDER BY created_at DESC
    ''', (user_id,))
    tokens = cursor.fetchall()
    conn.close()

    tokens_list = []
    for token in tokens:
        tokens_list.append({
            'id': token[0],
            'name': token[1],
            'token': token[2],
            'expires_at': token[3],
            'created_at': token[4]
        })

    return jsonify(tokens_list)


@app.route('/api/users/generate-token', methods=['POST'])
@login_required
def generate_user_token():
    data = request.json
    user_id = session.get('user_id')
    token_name = data.get('name')
    days = data.get('days', 30)

    if not token_name:
        return jsonify({'error': 'Token名称不能为空'}), 400

    # 生成随机Token
    token = secrets.token_urlsafe(32)

    # 计算过期时间
    expires_at = get_beijing_time() + timedelta(days=days)

    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute('''
        INSERT INTO user_tokens (user_id, token, name, expires_at)
        VALUES (?, ?, ?, ?)
    ''', (user_id, token, token_name, expires_at.isoformat()))

    token_id = cursor.lastrowid
    conn.commit()
    conn.close()

    return jsonify({
        'message': 'Token生成成功',
        'token': token,
        'expires_at': expires_at.isoformat(),
        'id': token_id
    })


@app.route('/api/users/tokens/<int:token_id>', methods=['DELETE'])
@login_required
def delete_user_token(token_id):
    user_id = session.get('user_id')
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute('DELETE FROM user_tokens WHERE id = ? AND user_id = ?', (token_id, user_id))
    conn.commit()
    conn.close()

    return jsonify({'message': 'Token删除成功'})


# 文件上传API
@app.route('/api/upload', methods=['POST'])
@login_required
def upload_file():
    if 'file' not in request.files:
        return jsonify({'error': '没有文件'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': '没有选择文件'}), 400

    if file and file.filename.endswith('.py'):
        # 保存文件到uploads目录
        uploads_dir = 'uploads'
        if not os.path.exists(uploads_dir):
            os.makedirs(uploads_dir)

        filepath = os.path.join(uploads_dir, file.filename)
        file.save(filepath)

        return jsonify({'message': '文件上传成功', 'filepath': filepath})

    return jsonify({'error': '只支持.py文件'}), 400


@app.route('/api/python-scripts', methods=['GET'])
def get_python_scripts():
    """获取所有Python脚本"""
    try:
        # 从uploads目录获取所有.py文件
        upload_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'uploads')
        if not os.path.exists(upload_dir):
            os.makedirs(upload_dir)

        python_scripts = []
        for filename in os.listdir(upload_dir):
            if filename.endswith('.py'):
                filepath = os.path.join(upload_dir, filename)
                script_info = {
                    'name': filename,
                    'path': filepath
                }
                python_scripts.append(script_info)

        return jsonify(python_scripts)
    except Exception as e:
        logging.error(f"获取Python脚本列表失败: {str(e)}")
        return jsonify({'error': f'获取Python脚本列表失败: {str(e)}'}), 500


# 修改数据插入API，移除手动Token验证，使用装饰器
@app.route('/api/database/insert', methods=['POST'])
@api_token_required
def insert_data_to_database():
    """
    通过API直接向数据库表插入数据
    支持会话认证和API Token认证
    """
    data = request.json
    logger.info(f"API call: insert_data_to_database with data: {data}")

    try:
        table_name = data.get('table_name')
        database_name = data.get('database_name')  # 从请求中获取数据库名
        records = data.get('records')  # 要插入的记录列表
        db_config_id = data.get('db_config_id')

        if not table_name or not database_name or not records or not db_config_id:  # 现在database_name是必需的
            logger.warning("Missing required parameters for database insert")
            return jsonify({'error': '缺少必要参数: table_name, database_name, records, db_config_id'}), 400

        # 获取数据库配置
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        cursor.execute('''
            SELECT db_type, host, port, username, password
            FROM db_configs 
            WHERE id = ?
        ''', (db_config_id,))

        db_config = cursor.fetchone()
        conn.close()

        if not db_config:
            logger.warning(f"Database configuration not found: {db_config_id}")
            return jsonify({'error': '数据库配置不存在'}), 404

        db_type, host, port, username, password = db_config

        # 使用传入的数据库名
        db_config_dict = {
            'db_type': db_type,
            'host': host,
            'port': port,
            'username': username,
            'password': password,
            'database': database_name  # 使用API调用中指定的数据库名
        }

        # 执行数据插入
        logger.info(f"Inserting {len(records)} records into table '{table_name}' in database '{database_name}'")
        success, result = insert_records_to_database(db_config_dict, table_name, records)
        if success:
            logger.info(
                f"Successfully inserted {len(records)} records into table '{table_name}' in database '{database_name}'")
            return jsonify({'message': f'成功插入 {len(records)} 条记录到表 {table_name}', 'result': result})
        else:
            logger.error(f"Failed to insert records into table '{table_name}' in database '{database_name}': {result}")
            return jsonify({'error': result}), 500

    except Exception as e:
        logger.error(f"Exception during database insert: {e}")
        return jsonify({'error': f'插入数据时出错: {str(e)}'}), 500


def insert_records_to_database(db_config, table_name, records):
    """
    向指定数据库表插入记录

    Args:
        db_config (dict): 数据库配置
        table_name (str): 表名
        records (list): 要插入的记录列表，每个记录是一个字典

    Returns:
        tuple: (success, result_or_error_message)
    """
    db_type = db_config['db_type']
    logger.info(f"Inserting records into {db_type} database, table: {table_name}")

    try:
        if db_type == 'mysql':
            # 连接到MySQL数据库
            logger.info(f"Connecting to MySQL database: {db_config['host']}:{db_config['port']}")
            connection = pymysql.connect(
                host=db_config['host'],
                port=db_config['port'] or 3306,
                user=db_config['username'],
                password=db_config['password'],
                database=db_config['database'],
                charset='utf8mb4',
                autocommit=True,
                cursorclass=pymysql.cursors.DictCursor
            )

            with connection:
                with connection.cursor() as cursor:
                    # 构造INSERT语句
                    if records:
                        columns = list(records[0].keys())
                        placeholders = ', '.join(['%s'] * len(columns))
                        columns_str = ', '.join([f'`{col}`' for col in columns])
                        sql = f"INSERT INTO `{table_name}` ({columns_str}) VALUES ({placeholders})"
                        logger.debug(f"MySQL insert SQL: {sql}")

                        # 批量插入数据
                        data = [[record[col] for col in columns] for record in records]
                        cursor.executemany(sql, data)
                        affected_rows = cursor.rowcount
                        logger.info(f"Inserted {affected_rows} rows into MySQL table {table_name}")

                return True, {'affected_rows': affected_rows}

        elif db_type == 'postgresql':
            # 连接到PostgreSQL数据库
            if psycopg2 is None:
                logger.error("PostgreSQL driver not installed")
                return False, "PostgreSQL驱动未安装"

            logger.info(f"Connecting to PostgreSQL database: {db_config['host']}:{db_config['port']}")
            connection = psycopg2.connect(
                host=db_config['host'],
                port=db_config['port'] or 5432,
                user=db_config['username'],
                password=db_config['password'],
                database=db_config['database']
            )

            with connection:
                with connection.cursor() as cursor:
                    # 构造INSERT语句
                    if records:
                        columns = list(records[0].keys())
                        placeholders = ', '.join(['%s'] * len(columns))
                        columns_str = ', '.join([f'"{col}"' for col in columns])
                        sql = f'INSERT INTO "{table_name}" ({columns_str}) VALUES ({placeholders})'
                        logger.debug(f"PostgreSQL insert SQL: {sql}")

                        # 批量插入数据
                        data = [[record[col] for col in columns] for record in records]
                        cursor.executemany(sql, data)
                        affected_rows = cursor.rowcount
                        connection.commit()
                        logger.info(f"Inserted {affected_rows} rows into PostgreSQL table {table_name}")

                return True, {'affected_rows': affected_rows}

        elif db_type == 'sqlite':
            # 连接到SQLite数据库
            logger.info(f"Connecting to SQLite database: {db_config['database']}")
            connection = sqlite3.connect(db_config['database'])

            with connection:
                cursor = connection.cursor()
                # 构造INSERT语句
                if records:
                    columns = list(records[0].keys())
                    placeholders = ', '.join(['?'] * len(columns))
                    columns_str = ', '.join([f'"{col}"' for col in columns])
                    sql = f'INSERT INTO "{table_name}" ({columns_str}) VALUES ({placeholders})'
                    logger.debug(f"SQLite insert SQL: {sql}")

                    # 批量插入数据
                    data = [[record[col] for col in columns] for record in records]
                    cursor.executemany(sql, data)
                    affected_rows = cursor.rowcount
                    logger.info(f"Inserted {affected_rows} rows into SQLite table {table_name}")

                return True, {'affected_rows': affected_rows}

        else:
            logger.error(f"Unsupported database type: {db_type}")
            return False, f"不支持的数据库类型: {db_type}"

    except Exception as e:
        logger.error(f"Error inserting records into {db_type} database table {table_name}: {e}")
        return False, f"插入数据时出错: {str(e)}"


# 添加一个临时的调试端点，用于手动添加Token（仅用于测试）
@app.route('/api/debug/add-token', methods=['POST'])
def debug_add_token():
    """
    调试端点：手动添加Token（仅用于测试环境）
    """
    data = request.json
    logger.info(f"API call: debug_add_token with data: {data}")
    user_id = data.get('user_id', 1)  # 默认用户ID为1
    token = data.get('token', '3TrEHpjtwcMGkCNVocj2w4ODNQquHyR_4-8kuXoAtNo')
    name = data.get('name', 'Debug Token')
    days = data.get('days', 30)

    # 计算过期时间（使用北京时间）
    expires_at = get_beijing_time() + timedelta(days=days)

    # 保存到数据库
    try:
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        cursor.execute('''
            INSERT OR REPLACE INTO user_tokens (user_id, token, name, expires_at)
            VALUES (?, ?, ?, ?)
        ''', (user_id, token, name, expires_at.isoformat()))

        conn.commit()
        conn.close()

        logger.info(f"Debug token added/updated successfully: {token}")
        return jsonify({
            'message': 'Token添加成功',
            'token': token,
            'expires_at': expires_at.isoformat()
        })
    except Exception as e:
        logger.error(f"Failed to add debug token: {e}")
        return jsonify({'error': f'Token添加失败: {str(e)}'}), 500

def load_email_config_from_db():
    """从数据库加载邮件配置到Flask应用配置"""
    try:
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        
        # 查询默认邮件配置
        cursor.execute('''
            SELECT mail_server, mail_port, mail_use_tls, mail_username, 
                   mail_password, mail_default_sender, mail_notification_enabled
            FROM email_configs
            WHERE is_default = 1
            ORDER BY id DESC
            LIMIT 1
        ''')
        config_row = cursor.fetchone()
        
        # 如果没有默认配置，获取第一个配置
        if not config_row:
            cursor.execute('''
                SELECT mail_server, mail_port, mail_use_tls, mail_username, 
                       mail_password, mail_default_sender, mail_notification_enabled
                FROM email_configs
                ORDER BY id ASC
                LIMIT 1
            ''')
            config_row = cursor.fetchone()
        
        conn.close()
        
        if config_row:
            # 如果存在配置，更新Flask应用配置
            app.config['MAIL_SERVER'] = config_row[0]
            app.config['MAIL_PORT'] = config_row[1]
            app.config['MAIL_USE_TLS'] = bool(config_row[2])
            app.config['MAIL_USERNAME'] = config_row[3]
            app.config['MAIL_PASSWORD'] = config_row[4]
            app.config['MAIL_DEFAULT_SENDER'] = config_row[5]
            app.config['MAIL_NOTIFICATION_ENABLED'] = bool(config_row[6])
            
            # 重新初始化邮件对象
            global mail
            mail.init_app(app)
            
            logger.info("邮件配置已从数据库加载")
            return True
        else:
            logger.info("数据库中没有找到邮件配置")
            return False
    except Exception as e:
        logger.error(f"从数据库加载邮件配置失败: {e}")
        return False


@app.route('/api/tasks')
@login_required
def get_tasks():
    logger.info("API call: get_tasks")
    
    # 获取筛选参数
    task_name_filter = request.args.get('task_name', '').strip()
    task_type_filter = request.args.get('task_type', '').strip()
    is_active_filter = request.args.get('is_active', '').strip()
    latest_status_filter = request.args.get('latest_status', '').strip()
    
    # 记录筛选参数
    logger.info(f"筛选参数 - task_name: {task_name_filter}, task_type: {task_type_filter}, is_active: {is_active_filter}, latest_status: {latest_status_filter}")
    
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    
    # 构建查询条件
    where_conditions = []
    params = []
    
    if task_name_filter:
        where_conditions.append("t.name LIKE ?")
        params.append(f"%{task_name_filter}%")
    
    if task_type_filter:
        where_conditions.append("t.task_type = ?")
        params.append(task_type_filter)
    
    if is_active_filter:
        if is_active_filter == 'true':
            where_conditions.append("t.is_active = 1")
        elif is_active_filter == 'false':
            where_conditions.append("t.is_active = 0")
    
    # 构建基础查询
    base_query = "SELECT t.* FROM tasks t"
    
    # 如果有执行状态筛选，需要关联task_logs表
    if latest_status_filter:
        # 使用子查询获取每个任务的最新执行状态
        base_query = """
            SELECT t.* FROM tasks t
            LEFT JOIN (
                SELECT DISTINCT task_id, status
                FROM task_logs tl1
                WHERE tl1.execution_time = (
                    SELECT MAX(tl2.execution_time)
                    FROM task_logs tl2
                    WHERE tl2.task_id = tl1.task_id
                )
            ) tl ON t.id = tl.task_id
        """
        if latest_status_filter == 'success':
            where_conditions.append("tl.status = 'success'")
        elif latest_status_filter == 'failed':
            where_conditions.append("tl.status = 'failed'")
        elif latest_status_filter == 'none':
            where_conditions.append("tl.task_id IS NULL")
    
    # 添加WHERE子句
    if where_conditions:
        base_query += " WHERE " + " AND ".join(where_conditions)
    
    # 添加排序和去重（当关联task_logs时可能有重复记录）
    if latest_status_filter:
        base_query += " GROUP BY t.id"
    base_query += " ORDER BY t.id DESC"
    
    # 记录最终查询
    logger.info(f"最终SQL查询: {base_query}")
    logger.info(f"查询参数: {params}")
    
    cursor.execute(base_query, params)
    tasks = cursor.fetchall()
    
    # 转换为字典列表
    tasks_list = []
    for task in tasks:
        # 解析依赖关系
        dependencies = []
        if len(task) > 9 and task[9]:  # dependencies字段
            try:
                dependencies = [int(dep) for dep in task[9].split(',') if dep]
            except ValueError:
                dependencies = []
        
        # 获取SQL脚本名称（如果适用）
        sql_script_name = None
        if task[2] == 'sql' and task[4]:  # task_type是sql且有sql_script_id
            cursor.execute('SELECT name FROM sql_scripts WHERE id = ?', (task[4],))
            sql_script = cursor.fetchone()
            if sql_script:
                sql_script_name = sql_script[0]
        
        # 获取最新执行状态
        latest_status = None
        cursor.execute('''
            SELECT status 
            FROM task_logs
            WHERE task_id = ?
            ORDER BY execution_time DESC
            LIMIT 1
        ''', (task[0],))
        status_result = cursor.fetchone()
        if status_result:
            latest_status = status_result[0]
        
        # 处理Python脚本路径，确保返回绝对路径
        script_path = task[3]
        if task[2] == 'python' and script_path:
            # 如果是相对路径，转换为绝对路径
            if not os.path.isabs(script_path):
                project_root = os.path.dirname(os.path.abspath(__file__))
                script_path = os.path.join(project_root, script_path)
                script_path = os.path.abspath(script_path)
        
        tasks_list.append({
            'id': task[0],
            'name': task[1],
            'task_type': task[2],
            'script_path': script_path,
            'sql_script_id': task[4],
            'sql_script_name': sql_script_name,
            'schedule_interval': task[5],
            'last_run': task[6],
            'next_run': task[7],
            'is_active': bool(task[8]),
            'dependencies': dependencies,
            'max_retries': task[10] if len(task) > 10 else 0,  # max_retries字段
            'retry_delay': task[11] if len(task) > 11 else 60,  # retry_delay字段
            'cron_expression': task[12] if len(task) > 12 else None,  # cron_expression字段
            'latest_status': latest_status  # 最新执行状态
        })
    
    conn.close()
    
    # 记录筛选条件和结果
    filter_info = []
    if task_name_filter:
        filter_info.append(f"任务名称包含: {task_name_filter}")
    if task_type_filter:
        filter_info.append(f"任务类型: {task_type_filter}")
    if is_active_filter:
        filter_info.append(f"启用状态: {is_active_filter}")
    if latest_status_filter:
        filter_info.append(f"执行状态: {latest_status_filter}")
    
    filter_str = ", ".join(filter_info) if filter_info else "无筛选条件"
    logger.info(f"Retrieved {len(tasks_list)} tasks with filters: {filter_str}")
    
    return jsonify({'tasks': tasks_list})

@app.route('/api/tasks/<int:task_id>', methods=['GET'])
@login_required
def get_task(task_id):
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM tasks WHERE id = ?', (task_id,))
    task = cursor.fetchone()

    if not task:
        conn.close()
        return jsonify({'error': '任务不存在'}), 404

    # 解析依赖关系
    dependencies = []
    if len(task) > 9 and task[9]:  # dependencies字段
        try:
            dependencies = [int(dep) for dep in task[9].split(',') if dep]
        except ValueError:
            dependencies = []

    # 获取SQL脚本名称（如果适用）
    sql_script_name = None
    if task[2] == 'sql' and task[4]:  # task_type是sql且有sql_script_id
        cursor.execute('SELECT name FROM sql_scripts WHERE id = ?', (task[4],))
        sql_script = cursor.fetchone()
        if sql_script:
            sql_script_name = sql_script[0]
    
    # 处理Python脚本路径，确保返回绝对路径
    script_path = task[3]
    if task[2] == 'python' and script_path:
        # 如果是相对路径，转换为绝对路径
        if not os.path.isabs(script_path):
            project_root = os.path.dirname(os.path.abspath(__file__))
            script_path = os.path.join(project_root, script_path)
            script_path = os.path.abspath(script_path)

    # 转换为字典
    task_dict = {
        'id': task[0],
        'name': task[1],
        'task_type': task[2],
        'script_path': script_path,
        'sql_script_id': task[4],
        'sql_script_name': sql_script_name,
        'schedule_interval': task[5],
        'last_run': task[6],
        'next_run': task[7],
        'is_active': bool(task[8]),
        'dependencies': dependencies,
        'max_retries': task[10] if len(task) > 10 else 0,  # max_retries字段
        'retry_delay': task[11] if len(task) > 11 else 60,  # retry_delay字段
        'cron_expression': task[12] if len(task) > 12 else None  # cron_expression字段
    }

    conn.close()
    return jsonify(task_dict)


@app.route('/api/tasks', methods=['POST'])
@login_required
def create_task():
    data = request.json
    logger.info(f"Creating task with data: {data}")

    # 验证必要字段
    if 'name' not in data or 'task_type' not in data:
        return jsonify({'error': '缺少必要字段'}), 400

    # 处理脚本路径，使其适应当前环境
    script_path = data.get('script_path')
    if script_path:
        script_path = normalize_script_path_for_storage(script_path)

    # 处理依赖关系
    dependencies = data.get('dependencies', [])
    dependencies_str = ','.join(map(str, dependencies)) if dependencies else None

    # 获取重试参数
    max_retries = data.get('max_retries', 0)
    retry_delay = data.get('retry_delay', 60)

    # 获取cron表达式
    cron_expression = data.get('cron_expression')

    # 计算调度间隔（如果提供了cron表达式，则为0）
    schedule_interval = 0 if cron_expression else data.get('schedule_interval', 0)

    # 计算下次运行时间
    now = get_beijing_time()
    next_run = now

    # 如果有cron表达式，使用cron表达式计算下次运行时间
    if cron_expression:
        try:
            cron = croniter(cron_expression, now)
            next_run = cron.get_next(datetime)
        except Exception as e:
            logger.error(f"Invalid cron expression when creating task: {cron_expression}, error: {e}")
            return jsonify({'error': f'无效的Cron表达式: {str(e)}'}), 400
    else:
        # 如果没有cron表达式，使用间隔时间计算下次运行时间
        next_run = now + timedelta(seconds=schedule_interval)

    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()

    cursor.execute('''
        INSERT INTO tasks (name, task_type, script_path, sql_script_id, schedule_interval, dependencies, max_retries, retry_delay, cron_expression, next_run)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    ''', (
        data['name'],
        data['task_type'],
        script_path,
        data.get('sql_script_id'),
        schedule_interval,
        dependencies_str,
        max_retries,
        retry_delay,
        cron_expression,
        next_run.isoformat()
    ))

    task_id = cursor.lastrowid
    conn.commit()
    conn.close()

    logger.info(f"Task created successfully with ID: {task_id}")
    return jsonify({'id': task_id, 'message': '任务创建成功'})


@app.route('/api/tasks/<int:task_id>', methods=['PUT'])
@login_required
def update_task(task_id):
    data = request.json
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()

    # 处理脚本路径，使其适应当前环境
    script_path = data.get('script_path')
    if script_path:
        script_path = normalize_script_path_for_storage(script_path)

    # 处理依赖关系
    dependencies = data.get('dependencies', [])
    dependencies_str = ','.join(map(str, dependencies)) if dependencies else None

    # 获取重试参数
    max_retries = data.get('max_retries', 0)
    retry_delay = data.get('retry_delay', 60)

    # 获取cron表达式
    cron_expression = data.get('cron_expression')

    # 如果有cron表达式，使用cron表达式计算下次运行时间
    next_run = None
    if cron_expression:
        try:
            now = get_beijing_time()
            cron = croniter(cron_expression, now)
            next_run = cron.get_next(datetime)
        except Exception as e:
            logger.error(f"Invalid cron expression when updating task {task_id}: {cron_expression}, error: {e}")
            # 如果cron表达式无效，不更新next_run字段

    if next_run:
        cursor.execute('''
            UPDATE tasks 
            SET name = ?, task_type = ?, script_path = ?, sql_script_id = ?, schedule_interval = ?, dependencies = ?, max_retries = ?, retry_delay = ?, cron_expression = ?, next_run = ?
            WHERE id = ?
        ''', (
            data['name'],
            data['task_type'],
            script_path,
            data.get('sql_script_id'),
            0 if cron_expression else data.get('schedule_interval', 0),
            dependencies_str,
            max_retries,
            retry_delay,
            cron_expression,
            next_run.isoformat(),
            task_id
        ))
    else:
        cursor.execute('''
            UPDATE tasks 
            SET name = ?, task_type = ?, script_path = ?, sql_script_id = ?, schedule_interval = ?, dependencies = ?, max_retries = ?, retry_delay = ?, cron_expression = ?
            WHERE id = ?
        ''', (
            data['name'],
            data['task_type'],
            script_path,
            data.get('sql_script_id'),
            0 if cron_expression else data.get('schedule_interval', 0),
            dependencies_str,
            max_retries,
            retry_delay,
            cron_expression,
            task_id
        ))

    conn.commit()
    conn.close()

    return jsonify({'message': '任务更新成功'})

    conn.commit()
    conn.close()

    return jsonify({'message': '任务更新成功'})


@app.route('/api/tasks/<int:task_id>/toggle', methods=['POST'])
@login_required
def toggle_task(task_id):
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()

    # 切换任务状态
    cursor.execute('''
        UPDATE tasks 
        SET is_active = NOT is_active 
        WHERE id = ?
    ''', (task_id,))

    conn.commit()
    conn.close()

    return jsonify({'message': '任务状态已更新'})


@app.route('/api/tasks/<int:task_id>', methods=['DELETE'])
@login_required
def delete_task(task_id):
    try:
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()

        # 先检查任务是否存在
        cursor.execute('SELECT id FROM tasks WHERE id = ?', (task_id,))
        task = cursor.fetchone()

        if not task:
            conn.close()
            return jsonify({'error': '任务不存在'}), 404

        # 删除任务
        cursor.execute('DELETE FROM tasks WHERE id = ?', (task_id,))

        conn.commit()
        conn.close()

        return jsonify({'message': '任务已删除'})
    except Exception as e:

        logging.error(f"删除任务失败: {str(e)}")
        return jsonify({'error': f'删除任务失败: {str(e)}'}), 500


@app.route('/api/tasks/<int:task_id>/run', methods=['POST'])
@login_required
def run_task_now(task_id):
    logger.info(f"API call: run_task_now for task ID {task_id}")
    try:
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()

        cursor.execute('''
            SELECT task_type, script_path, sql_script_id 
            FROM tasks 
            WHERE id = ?
        ''', (task_id,))
        task = cursor.fetchone()

        if not task:
            conn.close()
            logger.warning(f"Task not found: {task_id}")
            return jsonify({'error': '任务不存在'}), 404

        task_type, script_path, sql_script_id = task

        # 根据任务类型执行相应的操作
        message = ''
        if task_type == 'python' and script_path:
            if not os.path.exists(script_path):
                conn.close()
                logger.warning(f"Script file not found: {script_path}")
                return jsonify({'error': '脚本文件不存在'}), 404

            # 执行脚本
            result = execute_script(script_path)
            # 记录执行日志
            if result and result.returncode == 0:
                log_content = f'Python脚本执行成功: {script_path}\n'
                if result.stdout:
                    log_content += f'输出:\n{result.stdout}'
                log_task_execution(task_id, 'success', log_content)
                message = '任务执行成功'
            else:
                error_msg = result.stderr if result else '未知错误'
                output_msg = result.stdout if result else ''
                log_content = f'Python脚本执行失败: {script_path}\n'
                if error_msg:
                    log_content += f'错误:\n{error_msg}\n'
                if output_msg:
                    log_content += f'输出:\n{output_msg}'
                log_task_execution(task_id, 'failed', log_content)
                message = f'任务执行失败: {error_msg}'

        elif task_type == 'sql' and sql_script_id:
            # 执行SQL脚本
            success, result = execute_sql_script(sql_script_id)
            # 记录执行日志
            if success:
                log_task_execution(task_id, 'success', f'SQL脚本执行成功: {sql_script_id}')
                send_task_notification(task_id, 'success', f'SQL脚本执行成功: {sql_script_id}')
                message = '任务执行成功'
            else:
                log_task_execution(task_id, 'failed', f'SQL脚本执行失败: {sql_script_id}, 错误: {result}')
                send_task_notification(task_id, 'failed', f'SQL脚本执行失败: {sql_script_id}, 错误: {result}')
                message = f'任务执行失败: {result}'
        else:
            conn.close()
            return jsonify({'error': '任务类型不支持或配置不完整'}), 400

        # 更新最后运行时间
        now = get_beijing_time()
        cursor.execute('''
            UPDATE tasks 
            SET last_run = ? 
            WHERE id = ?
        ''', (now.isoformat(), task_id))

        conn.commit()
        conn.close()

        logger.info(f"Task {task_id} executed with message: {message}")
        return jsonify({'message': message})
    except Exception as e:
        logger.error(f"Error in run_task_now for task {task_id}: {e}", exc_info=True)
        # 确保数据库连接被关闭
        if 'conn' in locals():
            try:
                conn.close()
            except:
                pass
        return jsonify({'error': f'执行任务时发生内部错误: {str(e)}'}), 500


@app.route('/api/tasks/<int:task_id>/latest-log', methods=['GET'])
@login_required
def get_task_latest_log(task_id):
    """获取任务最新执行日志"""
    logger.info(f"API call: get_task_latest_log for task ID {task_id}")
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()

    # 先尝试获取任务信息以确认任务存在
    cursor.execute('SELECT name, task_type, sql_script_id FROM tasks WHERE id = ?', (task_id,))
    task = cursor.fetchone()

    if not task:
        conn.close()
        logger.warning(f"Task not found: {task_id}")
        return jsonify({'message': '任务不存在'}), 404

    task_name, task_type, sql_script_id = task

    cursor.execute('''
        SELECT status, message, execution_time
        FROM task_logs
        WHERE task_id = ?
        ORDER BY execution_time DESC
        LIMIT 1
    ''', (task_id,))

    log = cursor.fetchone()
    conn.close()

    if not log:
        logger.info(f"No log found for task ID {task_id}")
        # 对于SQL脚本任务，提供更具体的提示
        if task_type == 'sql' and sql_script_id:
            return jsonify({'message': f'暂无执行日志（SQL脚本ID: {sql_script_id}）'})
        else:
            return jsonify({'message': '暂无执行日志'})

    status, message, execution_time = log
    logger.info(f"Retrieved latest log for task {task_id}: status={status}")

    # 对于SQL脚本任务，在日志中添加更多信息
    if task_type == 'sql' and sql_script_id:
        message = f'[SQL脚本ID: {sql_script_id}] {message}'

    return jsonify({
        'task_name': task_name,
        'status': status,
        'message': message,
        'execution_time': execution_time
    })


@app.route('/api/sql-alerts/<int:alert_id>/latest-log', methods=['GET'])
@login_required
def get_sql_alert_latest_log(alert_id):
    """获取SQL预警最新执行日志"""
    logger.info(f"API call: get_sql_alert_latest_log for alert ID {alert_id}")
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()

    # 先尝试获取预警信息以确认预警存在
    cursor.execute('SELECT name FROM sql_alerts WHERE id = ?', (alert_id,))
    alert = cursor.fetchone()

    if not alert:
        conn.close()
        logger.warning(f"SQL alert not found: {alert_id}")
        return jsonify({'message': 'SQL预警不存在'}), 404

    alert_name = alert[0]

    cursor.execute('''
        SELECT status, message, execution_time
        FROM sql_alert_logs
        WHERE alert_id = ?
        ORDER BY execution_time DESC
        LIMIT 1
    ''', (alert_id,))

    log = cursor.fetchone()
    conn.close()

    if not log:
        logger.info(f"No log found for SQL alert ID {alert_id}")
        return jsonify({'message': '暂无执行日志'})

    status, message, execution_time = log
    logger.info(f"Retrieved latest log for SQL alert {alert_id}: status={status}")

    return jsonify({
        'alert_name': alert_name,
        'status': status,
        'message': message,
        'execution_time': execution_time
    })


@app.route('/api/sql-alerts/<int:alert_id>/latest-log-details', methods=['GET'])
@login_required
def get_sql_alert_latest_log_details(alert_id):
    """获取SQL预警最新执行详细日志"""
    logger.info(f"API call: get_sql_alert_latest_log_details for alert ID {alert_id}")
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()

    # 先尝试获取预警信息以确认预警存在
    cursor.execute('SELECT name FROM sql_alerts WHERE id = ?', (alert_id,))
    alert = cursor.fetchone()

    if not alert:
        conn.close()
        logger.warning(f"SQL alert not found: {alert_id}")
        return jsonify({'message': 'SQL预警不存在'}), 404

    alert_name = alert[0]

    cursor.execute('''
        SELECT status, message, execution_time, details
        FROM sql_alert_logs
        WHERE alert_id = ? AND details IS NOT NULL AND details != ''
        ORDER BY execution_time DESC
        LIMIT 1
    ''', (alert_id,))

    log = cursor.fetchone()
    conn.close()

    if not log:
        logger.info(f"No log found for SQL alert ID {alert_id}")
        return jsonify({'message': '暂无执行日志'})

    status, message, execution_time, details = log
    logger.info(f"Retrieved latest detailed log for SQL alert {alert_id}: status={status}")

    return jsonify({
        'alert_name': alert_name,
        'status': status,
        'message': message,
        'execution_time': execution_time,
        'details': details
    })


@app.route('/api/tasks/<int:task_id>/logs', methods=['GET'])
@login_required
def get_task_logs(task_id):
    """获取任务的所有执行日志（支持分页）"""
    logger.info(f"API call: get_task_logs for task ID {task_id}")

    # 获取分页参数
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 10, type=int)

    # 限制每页最大记录数
    per_page = min(per_page, 100)

    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()

    # 先尝试获取任务信息以确认任务存在
    cursor.execute('SELECT task_type, sql_script_id FROM tasks WHERE id = ?', (task_id,))
    task = cursor.fetchone()

    if not task:
        conn.close()
        logger.warning(f"Task not found: {task_id}")
        return jsonify({'message': '任务不存在'}), 404

    # 获取总记录数
    cursor.execute('SELECT COUNT(*) FROM task_logs WHERE task_id = ?', (task_id,))
    total_logs = cursor.fetchone()[0]

    # 计算偏移量
    offset = (page - 1) * per_page

    # 获取日志记录
    cursor.execute('''
        SELECT status, message, execution_time
        FROM task_logs
        WHERE task_id = ?
        ORDER BY execution_time DESC
        LIMIT ? OFFSET ?
    ''', (task_id, per_page, offset))

    logs = cursor.fetchall()
    conn.close()

    # 转换为字典列表
    logs_list = []
    for log in logs:
        status, message, execution_time = log
        logs_list.append({
            'status': status,
            'message': message,
            'execution_time': execution_time
        })

    return jsonify({
        'logs': logs_list,
        'pagination': {
            'page': page,
            'per_page': per_page,
            'total': total_logs,
            'pages': (total_logs + per_page - 1) // per_page
        }
    })


@app.route('/api/notification-logs', methods=['GET'])
@login_required
def get_notification_logs():
    """获取预警日志记录（支持分页和筛选）"""
    try:
        logger.info("API call: get_notification_logs")
        logger.info(f"Session data: {dict(session)}")
        logger.info(f"Request headers: {dict(request.headers)}")
        logger.info(f"Request args: {dict(request.args)}")
        logger.info(f"Request full URL: {request.url}")
    except Exception as e:
        logger.error(f"Error logging request info: {str(e)}")
    
    try:
        # 获取分页参数
        page_str = request.args.get('page', '1')
        per_page_str = request.args.get('per_page', '20')
        task_id_str = request.args.get('task_id')
        task_name = request.args.get('task_name')
        alert_type = request.args.get('alert_type')
        status = request.args.get('status')
        
        # 转换参数类型
        try:
            page = int(page_str) if page_str else 1
            per_page = int(per_page_str) if per_page_str else 20
            task_id = int(task_id_str) if task_id_str else None
        except ValueError as e:
            logger.error(f"Parameter conversion error: {str(e)}")
            return jsonify({'error': f'Invalid parameter format: {str(e)}'}), 400
        
        # 参数验证
        if page < 1:
            logger.error(f"Invalid page parameter: {page}")
            return jsonify({'error': 'Invalid page parameter'}), 400
        
        if per_page < 1 or per_page > 100:
            logger.error(f"Invalid per_page parameter: {per_page}")
            return jsonify({'error': 'Invalid per_page parameter'}), 400
        
        logger.info(f"Request params: page={page}, per_page={per_page}, task_id={task_id}, task_name={task_name}, alert_type={alert_type}, status={status}")

        # 限制每页最大记录数
        per_page = min(per_page, 100)

        # 连接数据库
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()

        # 构建查询条件
        where_conditions = []
        params = []

        if task_id:
            where_conditions.append("nl.task_id = ?")
            params.append(task_id)

        if task_name:
            where_conditions.append("nl.task_name LIKE ?")
            params.append(f"%{task_name}%")

        if alert_type:
            where_conditions.append("nl.alert_type = ?")
            params.append(alert_type)

        if status:
            where_conditions.append("nl.status = ?")
            params.append(status)

        where_clause = "WHERE " + " AND ".join(where_conditions) if where_conditions else ""

        # 获取总记录数
        cursor.execute(f'''
            SELECT COUNT(*)
            FROM notification_logs nl
            {where_clause}
        ''', params)
        total_logs = cursor.fetchone()[0]

        # 计算偏移量
        offset = (page - 1) * per_page

        # 获取日志记录
        cursor.execute(f'''
            SELECT nl.id, nl.task_id, nl.task_name, nl.alert_type, 
                   ec.config_name, nl.recipients, nl.subject, nl.status, 
                   nl.error_message, nl.sent_time
            FROM notification_logs nl
            LEFT JOIN email_configs ec ON nl.email_config_id = ec.id
            {where_clause}
            ORDER BY nl.sent_time DESC
            LIMIT ? OFFSET ?
        ''', params + [per_page, offset])

        logs = cursor.fetchall()
        conn.close()
        
    except sqlite3.Error as e:
        logger.error(f"Database error: {str(e)}")
        return jsonify({'error': f'Database error: {str(e)}'}), 500
    except Exception as e:
        logger.error(f"Unexpected error: {str(e)}")
        return jsonify({'error': f'Unexpected error: {str(e)}'}), 500

    # 转换为字典列表
    logs_list = []
    for log in logs:
        log_id, task_id, task_name, alert_type, config_name, recipients, subject, status, error_message, sent_time = log
        
        logs_list.append({
            'id': log_id,
            'task_id': task_id,
            'task_name': task_name,
            'alert_type': alert_type,
            'config_name': config_name,
            'recipients': recipients,
            'subject': subject,
            'status': status,
            'error_message': error_message,
            'sent_time': sent_time
        })

    return jsonify({
        'logs': logs_list,
        'pagination': {
            'page': page,
            'per_page': per_page,
            'total': total_logs,
            'pages': (total_logs + per_page - 1) // per_page
        }
    })


@app.route('/api/notification-logs/<int:log_id>', methods=['GET'])
@login_required
def get_notification_log_detail(log_id):
    """获取单个预警日志详情"""
    try:
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        
        cursor.execute('''
            SELECT nl.id, nl.task_id, nl.task_name, nl.alert_type, 
                   ec.config_name, nl.recipients, nl.subject, nl.body, nl.status, 
                   nl.error_message, nl.sent_time
            FROM notification_logs nl
            LEFT JOIN email_configs ec ON nl.email_config_id = ec.id
            WHERE nl.id = ?
        ''', (log_id,))
        
        log = cursor.fetchone()
        conn.close()
        
        if not log:
            return jsonify({'error': '日志记录不存在'}), 404
        
        log_id, task_id, task_name, alert_type, config_name, recipients, subject, body, status, error_message, sent_time = log
        
        return jsonify({
            'log': {
                'id': log_id,
                'task_id': task_id,
                'task_name': task_name,
                'alert_type': alert_type,
                'config_name': config_name,
                'recipients': recipients,
                'subject': subject,
                'body': body,
                'status': status,
                'error_message': error_message,
                'sent_time': sent_time
            }
        })
    except Exception as e:
        logger.error(f"获取预警日志详情失败: {str(e)}")
        return jsonify({'error': '获取预警日志详情失败'}), 500


if __name__ == '__main__':
    logger.info("Starting web scheduler application")
    logger.info(f"Python version: {sys.version}")
    logger.info(f"Current working directory: {os.getcwd()}")
    logger.info(f"Database path: {DB_PATH}")
    logger.info("Initializing database...")
    init_db()
    logger.info("Database initialization completed")
    
    # 从数据库加载邮件配置
    logger.info("Loading email configuration from database...")
    load_email_config_from_db()
    logger.info("Email configuration loaded successfully")

    # 启动调度器线程
    scheduler_thread = threading.Thread(target=task_scheduler)
    scheduler_thread.daemon = True
    scheduler_thread.start()
    logger.info("Task scheduler thread started")

    # 尝试使用Waitress生产级WSGI服务器，提高稳定性
    try:
        logger.info("Trying to start with Waitress WSGI server...")
        from waitress import serve
        logger.info("Waitress imported successfully")
        logger.info("Starting server with Waitress on http://127.0.0.1:5000")
        logger.info(f"Waitress configuration: host='127.0.0.1', port=5000, threads=4")
        logger.info("Server is about to start serving...")
        serve(app, host='127.0.0.1', port=5000, threads=4)
        logger.info("Waitress server has stopped (this should not happen in normal operation)")
    except ImportError:
        logger.warning("Waitress not available, falling back to Flask development server")
        logger.info("Install Waitress with: pip install waitress")
        
        # 在Windows上，使用更稳定的配置
        import os
        if os.name == 'nt':  # Windows系统
            # 使用更稳定的配置，禁用调试模式和重载器
            logger.info("Starting Flask development server on Windows with stable configuration")
            logger.info("Flask configuration: debug=False, host='127.0.0.1', port=5000, use_reloader=False, threaded=True")
            app.run(debug=False, host='127.0.0.1', port=5000, use_reloader=False, threaded=True)
        else:
            logger.info("Starting Flask development server on non-Windows system")
            logger.info("Flask configuration: debug=True, host='127.0.0.1', port=5000, use_reloader=False")
            app.run(debug=True, host='127.0.0.1', port=5000, use_reloader=False)
    except Exception as e:
        logger.error(f"Failed to start app with Waitress: {e}")
        logger.info("Trying alternative startup method...")
        # 备用启动方式
        try:
            from werkzeug.serving import run_simple
            logger.info("Starting with werkzeug.run_simple...")
            logger.info("Configuration: hostname='127.0.0.1', port=5000, application=app, use_reloader=False, use_debugger=False")
            run_simple('127.0.0.1', 5000, app, use_reloader=False, use_debugger=False)
            logger.info("Werkzeug server has stopped (this should not happen in normal operation)")
        except Exception as e2:
            logger.error(f"Failed to start app with alternative method: {e2}")
            logger.error(traceback.format_exc())
            # 最后的备用方案，使用最基本的配置
            logger.info("Trying last resort startup method...")
            logger.info("Configuration: host='127.0.0.1', port=5000")
            app.run(host='127.0.0.1', port=5000)
            logger.info("Last resort server has stopped (this should not happen in normal operation)")